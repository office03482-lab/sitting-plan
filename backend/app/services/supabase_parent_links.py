"""Supabase-backed parent/guardian linking workflows."""

from __future__ import annotations

import io
import secrets
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook

from app.services.supabase_admin import create_supabase_admin_client


ACADEMIC_SCHEMA = "academic"
PARENT_PERMISSIONS = [
    "parent_intelligence",
    "parent_intelligence.view",
    "parent_intelligence.alerts",
    "parent_intelligence.communication",
    "parent_intelligence.reports",
    "edupay.parent_portal",
]


def _client():
    return create_supabase_admin_client()


def _public_table(name: str):
    return _client().table(name)


def _academic_table(name: str):
    return _client().schema(ACADEMIC_SCHEMA).table(name)


def _normalize(value: Any) -> str:
    return str(value or "").strip()


def _normalize_optional(value: Any) -> str | None:
    text = _normalize(value)
    return text or None


def _lower(value: Any) -> str:
    return _normalize(value).lower()


def _make_temp_password() -> str:
    return f"Aspire@{secrets.token_urlsafe(8)}"


def _make_fallback_parent_email(*, full_name: str, phone: str, school_id: str) -> str:
    slug = "".join(ch for ch in full_name.lower() if ch.isalnum())[:24] or "parent"
    digits = "".join(ch for ch in phone if ch.isdigit())[-6:] or secrets.token_hex(3)
    school_slug = "".join(ch for ch in school_id.lower() if ch.isalnum())[:8] or "school"
    return f"{slug}.{digits}.{school_slug}@parent.local"


def _get_student(school_id: str, student_id: str) -> dict[str, Any]:
    rows = list(
        _public_table("students")
        .select("id,school_id,profile_id,full_name,roll_number,guardian_name,guardian_phone,metadata")
        .eq("school_id", school_id)
        .eq("id", student_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Student not found")
    return dict(rows[0])


def _serialize_guardian(guardian: dict[str, Any], *, linked_student_count: int = 0, temporary_password: str | None = None) -> dict[str, Any]:
    return {
        "id": _normalize(guardian.get("id")),
        "school_id": _normalize(guardian.get("school_id")),
        "profile_id": _normalize_optional(guardian.get("profile_id")),
        "guardian_code": _normalize_optional(guardian.get("guardian_code")),
        "full_name": _normalize(guardian.get("full_name")),
        "relation_type": _normalize(guardian.get("relation_type")) or "parent",
        "email": _normalize_optional(guardian.get("email")),
        "phone": _normalize_optional(guardian.get("phone")),
        "address": guardian.get("address"),
        "is_active": bool(guardian.get("is_active", True)),
        "linked_student_count": linked_student_count,
        "has_login": bool(_normalize(guardian.get("profile_id"))),
        "temporary_password": temporary_password,
        "created_at": guardian.get("created_at"),
        "updated_at": guardian.get("updated_at"),
    }


def _serialize_student_guardian_link(link_row: dict[str, Any], guardian_row: dict[str, Any]) -> dict[str, Any]:
    payload = _serialize_guardian(guardian_row)
    payload.update(
        {
            "link_id": _normalize(link_row.get("id")),
            "student_id": _normalize(link_row.get("student_id")),
            "guardian_id": _normalize(link_row.get("guardian_id")),
            "is_primary": bool(link_row.get("is_primary", False)),
            "can_receive_notifications": bool(link_row.get("can_receive_notifications", True)),
            "linked_at": link_row.get("created_at"),
        }
    )
    return payload


def _list_guardian_links_for_student(school_id: str, student_id: str) -> list[dict[str, Any]]:
    link_rows = list(
        _academic_table("student_guardians")
        .select("*")
        .eq("school_id", school_id)
        .eq("student_id", student_id)
        .execute()
        .data
        or []
    )
    if not link_rows:
        return []
    guardian_ids = [_normalize(row.get("guardian_id")) for row in link_rows if _normalize(row.get("guardian_id"))]
    guardian_rows = list(
        _academic_table("guardians")
        .select("*")
        .eq("school_id", school_id)
        .in_("id", guardian_ids)
        .execute()
        .data
        or []
    )
    guardian_map = {_normalize(row.get("id")): dict(row) for row in guardian_rows}
    return [
        _serialize_student_guardian_link(dict(link_row), guardian_map.get(_normalize(link_row.get("guardian_id")), {}))
        for link_row in link_rows
        if guardian_map.get(_normalize(link_row.get("guardian_id")))
    ]


def _count_linked_students_by_guardian_ids(school_id: str, guardian_ids: list[str]) -> dict[str, int]:
    if not guardian_ids:
        return {}
    rows = list(
        _academic_table("student_guardians")
        .select("guardian_id,student_id")
        .eq("school_id", school_id)
        .in_("guardian_id", guardian_ids)
        .execute()
        .data
        or []
    )
    counts: dict[str, int] = {}
    for row in rows:
        guardian_id = _normalize(row.get("guardian_id"))
        if not guardian_id:
            continue
        counts[guardian_id] = counts.get(guardian_id, 0) + 1
    return counts


def _find_guardian_candidates(school_id: str, *, email: str | None = None, phone: str | None = None, full_name: str | None = None) -> list[dict[str, Any]]:
    rows = list(
        _academic_table("guardians")
        .select("*")
        .eq("school_id", school_id)
        .eq("is_active", True)
        .execute()
        .data
        or []
    )
    wanted_email = _lower(email)
    wanted_phone = _normalize(phone)
    wanted_name = _lower(full_name)
    matches: list[dict[str, Any]] = []
    for row in rows:
        if wanted_email and _lower(row.get("email")) == wanted_email:
            matches.append(dict(row))
            continue
        if wanted_phone and _normalize(row.get("phone")) == wanted_phone:
            matches.append(dict(row))
            continue
        if wanted_name and _lower(row.get("full_name")) == wanted_name:
            matches.append(dict(row))
    return matches


def _get_or_create_parent_profile(
    *,
    school_id: str,
    full_name: str,
    email: str | None,
    phone: str | None,
    relation_type: str,
    create_login: bool,
    password: str | None,
) -> tuple[str | None, str | None, str | None]:
    if not create_login:
        return None, _normalize_optional(email), None

    from app.routes.auth import _ensure_managed_role, normalize_permissions

    supplied_email = _normalize_optional(email)
    safe_email = supplied_email or _make_fallback_parent_email(full_name=full_name, phone=_normalize(phone), school_id=school_id)
    temp_password = password or _make_temp_password()

    profile_rows = list(
        _public_table("profiles")
        .select("id,email,full_name,display_name,is_active,metadata")
        .ilike("email", safe_email)
        .limit(1)
        .execute()
        .data
        or []
    )
    profile_id = _normalize_optional(profile_rows[0].get("id")) if profile_rows else None

    if not profile_id:
        try:
            user_response = _client().auth.admin.create_user(
                {
                    "email": safe_email,
                    "password": temp_password,
                    "email_confirm": True,
                    "user_metadata": {
                        "full_name": full_name,
                        "display_name": full_name,
                        "relation_type": relation_type,
                    },
                }
            )
        except Exception as exc:
            detail = _lower(exc)
            if "already" not in detail and "duplicate" not in detail:
                raise HTTPException(status_code=500, detail=str(exc).strip() or "Failed to create parent account") from exc
            retry_rows = list(
                _public_table("profiles")
                .select("id,email,full_name,display_name,is_active,metadata")
                .ilike("email", safe_email)
                .limit(1)
                .execute()
                .data
                or []
            )
            profile_id = _normalize_optional(retry_rows[0].get("id")) if retry_rows else None
        else:
            created_user = getattr(user_response, "user", None)
            profile_id = _normalize_optional(getattr(created_user, "id", None))

    if not profile_id:
        raise HTTPException(status_code=500, detail="Failed to resolve parent profile")

    _public_table("profiles").update(
        {
            "full_name": full_name,
            "display_name": full_name,
            "is_active": True,
            "metadata": {
                "user_type": "non_teaching",
                "relation_type": relation_type,
                "contact_phone": _normalize_optional(phone),
                "managed_by": "parent_linking",
            },
        }
    ).eq("id", profile_id).execute()

    role_row = _ensure_managed_role(
        school_id,
        profile_id,
        full_name=full_name,
        selected_role="parent",
        user_type="non_teaching",
        permissions=normalize_permissions(PARENT_PERMISSIONS),
        supabase=_client(),
    )
    membership_rows = list(
        _public_table("school_memberships")
        .select("id")
        .eq("school_id", school_id)
        .eq("profile_id", profile_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    if membership_rows:
        _public_table("school_memberships").update(
            {
                "role_id": role_row["id"],
                "status": "active",
                "is_active": True,
                "metadata": {"source": "parent_linking"},
            }
        ).eq("id", membership_rows[0]["id"]).execute()
        return profile_id, safe_email, None

    _public_table("school_memberships").insert(
        {
            "school_id": school_id,
            "profile_id": profile_id,
            "role_id": role_row["id"],
            "status": "active",
            "is_primary": False,
            "is_active": True,
            "metadata": {"source": "parent_linking"},
        }
    ).execute()
    return profile_id, safe_email, temp_password


def _sync_student_parent_metadata(school_id: str, student_id: str) -> None:
    link_rows = _list_guardian_links_for_student(school_id, student_id)
    guardian_profile_ids = [item["profile_id"] for item in link_rows if item.get("profile_id")]
    guardian_emails = [item["email"] for item in link_rows if item.get("email")]
    primary = next((item for item in link_rows if item.get("is_primary")), link_rows[0] if link_rows else None)

    student_rows = list(
        _public_table("students")
        .select("id,metadata")
        .eq("school_id", school_id)
        .eq("id", student_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not student_rows:
        return
    metadata = student_rows[0].get("metadata") or {}
    if not isinstance(metadata, dict):
        metadata = {}
    metadata = {
        **metadata,
        "parent_profile_id": primary.get("profile_id") if primary else None,
        "guardian_profile_id": primary.get("profile_id") if primary else None,
        "parent_profile_ids": guardian_profile_ids,
        "guardian_profile_ids": guardian_profile_ids,
        "parent_email": primary.get("email") if primary else None,
        "guardian_email": primary.get("email") if primary else None,
        "parent_emails": guardian_emails,
    }
    update_payload = {
        "metadata": metadata,
        "guardian_name": primary.get("full_name") if primary else None,
        "guardian_phone": primary.get("phone") if primary else None,
    }
    _public_table("students").update(update_payload).eq("school_id", school_id).eq("id", student_id).execute()


def list_parent_directory(school_id: str, *, search: str | None = None, limit: int = 100) -> list[dict[str, Any]]:
    rows = list(
        _academic_table("guardians")
        .select("*")
        .eq("school_id", school_id)
        .eq("is_active", True)
        .order("updated_at", desc=True)
        .limit(limit)
        .execute()
        .data
        or []
    )
    if search:
        wanted = _lower(search)
        rows = [
            dict(row)
            for row in rows
            if wanted in _lower(row.get("full_name"))
            or wanted in _lower(row.get("email"))
            or wanted in _lower(row.get("phone"))
        ]
    guardian_ids = [_normalize(row.get("id")) for row in rows if _normalize(row.get("id"))]
    counts = _count_linked_students_by_guardian_ids(school_id, guardian_ids)
    return [_serialize_guardian(dict(row), linked_student_count=counts.get(_normalize(row.get("id")), 0)) for row in rows]


def list_student_parents(school_id: str, student_id: str) -> list[dict[str, Any]]:
    _get_student(school_id, student_id)
    return _list_guardian_links_for_student(school_id, student_id)


def create_or_link_parent(
    school_id: str,
    student_id: str,
    *,
    guardian_id: str | None = None,
    full_name: str | None = None,
    email: str | None = None,
    phone: str | None = None,
    relation_type: str | None = None,
    address: str | None = None,
    is_primary: bool = False,
    can_receive_notifications: bool = True,
    create_login: bool = True,
    password: str | None = None,
) -> dict[str, Any]:
    student = _get_student(school_id, student_id)
    relation_value = _normalize(relation_type) or "parent"
    if relation_value not in {"parent", "father", "mother", "guardian", "sponsor"}:
        raise HTTPException(status_code=400, detail="Invalid relation_type")

    guardian_row: dict[str, Any] | None = None
    temporary_password: str | None = None
    resolved_email = _normalize_optional(email)

    if guardian_id:
        guardian_rows = list(
            _academic_table("guardians")
            .select("*")
            .eq("school_id", school_id)
            .eq("id", guardian_id)
            .limit(1)
            .execute()
            .data
            or []
        )
        if not guardian_rows:
            raise HTTPException(status_code=404, detail="Parent not found")
        guardian_row = dict(guardian_rows[0])
    else:
        normalized_name = _normalize(full_name)
        if not normalized_name:
            raise HTTPException(status_code=400, detail="Parent full_name is required")
        matches = _find_guardian_candidates(school_id, email=resolved_email, phone=phone, full_name=normalized_name)
        guardian_row = dict(matches[0]) if matches else None

        profile_id: str | None = None
        if create_login:
            profile_id, resolved_email, temporary_password = _get_or_create_parent_profile(
                school_id=school_id,
                full_name=normalized_name,
                email=resolved_email,
                phone=phone,
                relation_type=relation_value,
                create_login=create_login,
                password=password,
            )
        if guardian_row:
            guardian_update = {
                "profile_id": profile_id or guardian_row.get("profile_id"),
                "full_name": normalized_name,
                "relation_type": relation_value,
                "email": resolved_email or guardian_row.get("email"),
                "phone": _normalize_optional(phone) or guardian_row.get("phone"),
                "address": address if address is not None else guardian_row.get("address"),
                "is_active": True,
            }
            _academic_table("guardians").update(guardian_update).eq("id", guardian_row["id"]).execute()
        else:
            created = _academic_table("guardians").insert(
                {
                    "school_id": school_id,
                    "profile_id": profile_id,
                    "guardian_code": None,
                    "full_name": normalized_name,
                    "relation_type": relation_value,
                    "email": resolved_email,
                    "phone": _normalize_optional(phone),
                    "address": address,
                    "metadata": {
                        "source": "parent_linking",
                        "student_roll_number": _normalize(student.get("roll_number")),
                    },
                    "is_active": True,
                }
            ).execute()
            guardian_rows = list(created.data or [])
            if not guardian_rows:
                raise HTTPException(status_code=500, detail="Failed to create parent")
            guardian_row = dict(guardian_rows[0])

    if guardian_row is None:
        raise HTTPException(status_code=500, detail="Failed to resolve parent record")

    if is_primary:
        _academic_table("student_guardians").update({"is_primary": False}).eq("school_id", school_id).eq("student_id", student_id).execute()

    existing_links = list(
        _academic_table("student_guardians")
        .select("*")
        .eq("school_id", school_id)
        .eq("student_id", student_id)
        .eq("guardian_id", guardian_row["id"])
        .limit(1)
        .execute()
        .data
        or []
    )
    if existing_links:
        _academic_table("student_guardians").update(
            {
                "is_primary": is_primary or bool(existing_links[0].get("is_primary", False)),
                "can_receive_notifications": can_receive_notifications,
            }
        ).eq("id", existing_links[0]["id"]).execute()
    else:
        _academic_table("student_guardians").insert(
            {
                "school_id": school_id,
                "student_id": student_id,
                "guardian_id": guardian_row["id"],
                "is_primary": is_primary,
                "can_receive_notifications": can_receive_notifications,
            }
        ).execute()

    _sync_student_parent_metadata(school_id, student_id)
    links = _list_guardian_links_for_student(school_id, student_id)
    linked = next((item for item in links if _normalize(item.get("guardian_id")) == _normalize(guardian_row.get("id"))), None)
    if not linked:
        raise HTTPException(status_code=500, detail="Failed to link parent")
    linked["temporary_password"] = temporary_password
    return linked


def unlink_parent(school_id: str, student_id: str, guardian_id: str) -> dict[str, Any]:
    rows = list(
        _academic_table("student_guardians")
        .select("*")
        .eq("school_id", school_id)
        .eq("student_id", student_id)
        .eq("guardian_id", guardian_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Parent link not found")
    _academic_table("student_guardians").delete().eq("id", rows[0]["id"]).execute()
    _sync_student_parent_metadata(school_id, student_id)
    return {"message": "Parent unlinked successfully", "student_id": student_id, "guardian_id": guardian_id}


def import_parent_links_from_excel(school_id: str, file_bytes: bytes) -> dict[str, Any]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid Excel file for parent import") from exc
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Parent import sheet is empty")

    headers = [str(cell or "").strip().lower() for cell in rows[0]]
    expected = {
        "student name": "student_name",
        "student roll no": "student_roll_no",
        "parent name": "parent_name",
        "parent email": "parent_email",
        "parent phone": "parent_phone",
    }
    header_map: dict[int, str] = {}
    for index, header in enumerate(headers):
        if header in expected:
            header_map[index] = expected[header]
    required = {"student_roll_no", "parent_name"}
    if not required.issubset(set(header_map.values())):
        raise HTTPException(status_code=400, detail="Excel must include at least 'Student Roll No' and 'Parent Name' columns")

    created_count = 0
    linked_count = 0
    skipped_count = 0
    errors: list[dict[str, Any]] = []
    credentials: list[dict[str, str]] = []

    student_rows = list(
        _public_table("students")
        .select("id,school_id,full_name,roll_number")
        .eq("school_id", school_id)
        .eq("is_active", True)
        .execute()
        .data
        or []
    )
    students_by_roll = {_lower(row.get("roll_number")): dict(row) for row in student_rows if _normalize(row.get("roll_number"))}

    for row_index, values in enumerate(rows[1:], start=2):
        payload = {header_map[index]: _normalize(values[index]) for index in header_map if index < len(values)}
        if not any(payload.values()):
            continue
        student = students_by_roll.get(_lower(payload.get("student_roll_no")))
        if not student:
            skipped_count += 1
            errors.append({"row": row_index, "error": "Student roll number not found", "student_roll_no": payload.get("student_roll_no")})
            continue
        try:
            linked = create_or_link_parent(
                school_id,
                _normalize(student.get("id")),
                full_name=payload.get("parent_name"),
                email=payload.get("parent_email") or None,
                phone=payload.get("parent_phone") or None,
                relation_type="parent",
                is_primary=False,
                can_receive_notifications=True,
                create_login=True,
            )
        except Exception as exc:
            skipped_count += 1
            errors.append({"row": row_index, "error": str(exc), "student_roll_no": payload.get("student_roll_no")})
            continue
        linked_count += 1
        if linked.get("temporary_password"):
            created_count += 1
            credentials.append(
                {
                    "student_roll_no": payload.get("student_roll_no") or "",
                    "parent_name": payload.get("parent_name") or "",
                    "email": linked.get("email") or "",
                    "temporary_password": linked.get("temporary_password") or "",
                }
            )

    return {
        "created_count": created_count,
        "linked_count": linked_count,
        "skipped_count": skipped_count,
        "errors": errors,
        "credentials": credentials,
    }

"""Supabase-native online test repository and lifecycle helpers."""

from __future__ import annotations

from io import BytesIO
from datetime import datetime, timedelta, timezone
from typing import Any
from uuid import UUID

from fastapi import HTTPException
from openpyxl import load_workbook

from app.services.ai_provider import AIProviderError, generate_json
from app.services.supabase_admin import get_supabase_admin_client

ONLINE_TESTS_SCHEMA = "online_tests"
ONLINE_TESTS_MODULE_KEY = "online_tests"
_ONLINE_TESTS_PUBLIC_PREFIX = "online_test_"


def _client():
    return get_supabase_admin_client()


def _table(name: str):
    # Access via public.online_test_* views (PostgREST workaround for
    # non-exposed private schema).  Requires applying the
    # 20260614_055_online_tests_public_views.sql migration first.
    return _client().table(f"{_ONLINE_TESTS_PUBLIC_PREFIX}{name}")


def _question_bank_table():
    return _client().table("online_test_question_bank")


def _normalize(value: Any) -> str:
    return str(value or "").strip()


def _normalize_optional_uuid(value: Any) -> str | None:
    text = _normalize(value)
    if not text:
        return None
    try:
        return str(UUID(text))
    except (TypeError, ValueError, AttributeError) as exc:
        raise HTTPException(status_code=400, detail="Expected a valid UUID value") from exc


def _to_iso_datetime(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).isoformat()
    text = _normalize(value)
    return text or None


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize_json_object(value: Any) -> dict[str, Any]:
    return dict(value) if isinstance(value, dict) else {}


def _log_audit_entry(
    *,
    school_id: str | None,
    profile_id: str | None,
    action: str,
    entity_id: str | None = None,
    payload: dict[str, Any] | None = None,
) -> None:
    row: dict[str, Any] = {
        "school_id": _normalize_optional_uuid(school_id),
        "profile_id": _normalize_optional_uuid(profile_id),
        "action": action,
        "module_key": ONLINE_TESTS_MODULE_KEY,
        "payload": payload or {},
    }
    entity_uuid = _normalize_optional_uuid(entity_id)
    if entity_uuid:
        row["entity_id"] = entity_uuid
    _client().table("audit_logs").insert(row).execute()


def _serialize_section(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": _normalize(row.get("id")),
        "test_id": _normalize(row.get("test_id")),
        "school_id": _normalize(row.get("school_id")),
        "title": _normalize(row.get("title")),
        "description": row.get("description"),
        "display_order": int(row.get("display_order") or 1),
        "question_type": _normalize(row.get("question_type")) or "mixed",
        "marks_per_question": float(row.get("marks_per_question") or 0),
        "negative_marks": float(row.get("negative_marks") or 0),
        "question_count": int(row.get("question_count") or 0),
        "is_active": bool(row.get("is_active", True)),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


def _serialize_test(row: dict[str, Any], sections: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    return {
        "id": _normalize(row.get("id")),
        "school_id": _normalize(row.get("school_id")),
        "title": _normalize(row.get("title")),
        "description": row.get("description"),
        "instructions": row.get("instructions"),
        "test_code": row.get("test_code"),
        "subject_id": _normalize(row.get("subject_id")) or None,
        "batch_id": _normalize(row.get("batch_id")) or None,
        "test_type": _normalize(row.get("test_type")) or "objective",
        "delivery_mode": _normalize(row.get("delivery_mode")) or "scheduled",
        "status": _normalize(row.get("status")) or "draft",
        "duration_minutes": int(row.get("duration_minutes") or 0),
        "total_marks": float(row.get("total_marks") or 0),
        "pass_marks": float(row.get("pass_marks")) if row.get("pass_marks") is not None else None,
        "max_attempts": int(row.get("max_attempts") or 1),
        "shuffle_questions": bool(row.get("shuffle_questions", False)),
        "shuffle_options": bool(row.get("shuffle_options", False)),
        "show_result_immediately": bool(row.get("show_result_immediately", False)),
        "allow_review": bool(row.get("allow_review", True)),
        "starts_at": row.get("starts_at"),
        "ends_at": row.get("ends_at"),
        "published_at": row.get("published_at"),
        "metadata": _normalize_json_object(row.get("metadata")),
        "is_active": bool(row.get("is_active", True)),
        "sections": sections or [],
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


def _serialize_question(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": _normalize(row.get("id")),
        "school_id": _normalize(row.get("school_id")),
        "test_id": _normalize(row.get("test_id")),
        "section_id": _normalize(row.get("section_id")),
        "question_code": row.get("question_code"),
        "display_order": int(row.get("display_order") or 1),
        "question_type": _normalize(row.get("question_type")) or "single_choice",
        "difficulty_level": _normalize(row.get("difficulty_level")) or "medium",
        "prompt_text": _normalize(row.get("prompt_text")),
        "option_items": list(row.get("option_items") or []),
        "answer_key": _normalize_json_object(row.get("answer_key")),
        "explanation": row.get("explanation"),
        "marks": float(row.get("marks") or 0),
        "negative_marks": float(row.get("negative_marks") or 0),
        "metadata": _normalize_json_object(row.get("metadata")),
        "is_active": bool(row.get("is_active", True)),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


def _serialize_question_bank_item(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": _normalize(row.get("id")),
        "school_id": _normalize(row.get("school_id")),
        "subject": row.get("subject"),
        "chapter": row.get("chapter"),
        "topic": row.get("topic"),
        "question_type": _normalize(row.get("question_type")) or "single_choice",
        "difficulty_level": _normalize(row.get("difficulty_level")) or "medium",
        "prompt_text": _normalize(row.get("prompt_text")),
        "option_items": list(row.get("option_items") or []),
        "answer_key": _normalize_json_object(row.get("answer_key")),
        "explanation": row.get("explanation"),
        "marks": float(row.get("marks") or 0),
        "negative_marks": float(row.get("negative_marks") or 0),
        "metadata": _normalize_json_object(row.get("metadata")),
        "is_active": bool(row.get("is_active", True)),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


def _sanitize_question_for_student(question: dict[str, Any]) -> dict[str, Any]:
    sanitized = dict(question)
    sanitized["answer_key"] = {}
    sanitized["explanation"] = None
    return sanitized


def _serialize_response(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": _normalize(row.get("id")),
        "school_id": _normalize(row.get("school_id")),
        "attempt_id": _normalize(row.get("attempt_id")),
        "test_id": _normalize(row.get("test_id")),
        "question_id": _normalize(row.get("question_id")),
        "student_id": _normalize(row.get("student_id")),
        "response_payload": _normalize_json_object(row.get("response_payload")),
        "is_marked_for_review": bool(row.get("is_marked_for_review", False)),
        "is_correct": row.get("is_correct"),
        "marks_awarded": float(row.get("marks_awarded")) if row.get("marks_awarded") is not None else None,
        "answered_at": row.get("answered_at"),
        "evaluated_at": row.get("evaluated_at"),
        "metadata": _normalize_json_object(row.get("metadata")),
        "is_active": bool(row.get("is_active", True)),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


def _serialize_attempt(row: dict[str, Any], responses: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    return {
        "id": _normalize(row.get("id")),
        "school_id": _normalize(row.get("school_id")),
        "test_id": _normalize(row.get("test_id")),
        "student_id": _normalize(row.get("student_id")),
        "attempt_number": int(row.get("attempt_number") or 1),
        "status": _normalize(row.get("status")) or "in_progress",
        "started_at": row.get("started_at"),
        "submitted_at": row.get("submitted_at"),
        "auto_submitted_at": row.get("auto_submitted_at"),
        "evaluated_at": row.get("evaluated_at"),
        "total_questions_snapshot": int(row.get("total_questions_snapshot") or 0),
        "answered_questions_snapshot": int(row.get("answered_questions_snapshot") or 0),
        "time_spent_seconds": int(row.get("time_spent_seconds") or 0),
        "metadata": _normalize_json_object(row.get("metadata")),
        "is_active": bool(row.get("is_active", True)),
        "responses": responses or [],
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


def _serialize_result(row: dict[str, Any]) -> dict[str, Any]:
    metadata = _normalize_json_object(row.get("metadata"))
    return {
        "id": _normalize(row.get("id")),
        "school_id": _normalize(row.get("school_id")),
        "attempt_id": _normalize(row.get("attempt_id")),
        "test_id": _normalize(row.get("test_id")),
        "student_id": _normalize(row.get("student_id")),
        "status": _normalize(row.get("status")) or "evaluated",
        "total_questions": int(row.get("total_questions") or 0),
        "attempted_questions": int(row.get("attempted_questions") or 0),
        "correct_answers": int(row.get("correct_answers") or 0),
        "incorrect_answers": int(row.get("incorrect_answers") or 0),
        "unanswered_questions": int(row.get("unanswered_questions") or 0),
        "score_obtained": float(row.get("score_obtained") or 0),
        "max_score": float(row.get("max_score") or 0),
        "percentage": float(row.get("percentage")) if row.get("percentage") is not None else None,
        "rank_in_batch": int(row.get("rank_in_batch")) if row.get("rank_in_batch") is not None else None,
        "rank_in_school": int(row.get("rank_in_school")) if row.get("rank_in_school") is not None else None,
        "passed": metadata.get("passed") if "passed" in metadata else None,
        "pass_marks": float(metadata.get("pass_marks")) if metadata.get("pass_marks") is not None else None,
        "published_at": row.get("published_at"),
        "metadata": metadata,
        "is_active": bool(row.get("is_active", True)),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
    }


def _load_sections_map(school_id: str, test_ids: list[str]) -> dict[str, list[dict[str, Any]]]:
    if not test_ids:
        return {}
    rows = list(
        _table("test_sections")
        .select("*")
        .eq("school_id", school_id)
        .in_("test_id", test_ids)
        .is_("deleted_at", "null")
        .order("display_order", desc=False)
        .execute()
        .data
        or []
    )
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        test_id = _normalize(row.get("test_id"))
        grouped.setdefault(test_id, []).append(_serialize_section(dict(row)))
    return grouped


def _get_test_row(school_id: str, test_id: str) -> dict[str, Any]:
    rows = list(
        _table("tests")
        .select("*")
        .eq("school_id", school_id)
        .eq("id", test_id)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Online test not found")
    return dict(rows[0])


def _get_attempt_row(school_id: str, attempt_id: str) -> dict[str, Any]:
    rows = list(
        _table("test_attempts")
        .select("*")
        .eq("school_id", school_id)
        .eq("id", attempt_id)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Test attempt not found")
    return dict(rows[0])


def _get_result_row(school_id: str, result_id: str) -> dict[str, Any]:
    rows = list(
        _table("test_results")
        .select("*")
        .eq("school_id", school_id)
        .eq("id", result_id)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Test result not found")
    return dict(rows[0])


def _get_student_by_profile_id(school_id: str, profile_id: str) -> dict[str, Any]:
    rows = list(
        _client()
        .table("students")
        .select("id, school_id, profile_id, batch_id, full_name")
        .eq("school_id", school_id)
        .eq("profile_id", profile_id)
        .eq("is_active", True)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        raise HTTPException(status_code=403, detail="Student profile is not linked to an active student record")
    return dict(rows[0])


def _ensure_default_section(school_id: str, test_id: str) -> dict[str, Any]:
    existing_rows = list(
        _table("test_sections")
        .select("*")
        .eq("school_id", school_id)
        .eq("test_id", test_id)
        .is_("deleted_at", "null")
        .order("display_order", desc=False)
        .limit(1)
        .execute()
        .data
        or []
    )
    if existing_rows:
        return dict(existing_rows[0])
    created = (
        _table("test_sections")
        .insert(
            {
                "school_id": school_id,
                "test_id": test_id,
                "section_code": "SEC-1",
                "title": "Section 1",
                "display_order": 1,
                "question_type": "mixed",
                "marks_per_question": 1,
                "negative_marks": 0,
                "question_count": 0,
            }
        )
        .execute()
    )
    rows = list(created.data or [])
    if not rows:
        raise HTTPException(status_code=500, detail="Default test section could not be created")
    return dict(rows[0])


def _resolve_section_id(school_id: str, test_id: str, section_id: str | None) -> str:
    candidate = _normalize(section_id)
    if candidate:
        rows = list(
            _table("test_sections")
            .select("id")
            .eq("school_id", school_id)
            .eq("test_id", test_id)
            .eq("id", candidate)
            .is_("deleted_at", "null")
            .limit(1)
            .execute()
            .data
            or []
        )
        if not rows:
            raise HTTPException(status_code=404, detail="Test section not found")
        return candidate

    sections = list(
        _table("test_sections")
        .select("id")
        .eq("school_id", school_id)
        .eq("test_id", test_id)
        .is_("deleted_at", "null")
        .order("display_order", desc=False)
        .execute()
        .data
        or []
    )
    if len(sections) == 1:
        return _normalize(sections[0].get("id"))
    if not sections:
        return _normalize(_ensure_default_section(school_id, test_id).get("id"))
    raise HTTPException(status_code=400, detail="section_id is required when a test has multiple sections")


def _recount_section_questions(school_id: str, section_id: str) -> None:
    response = (
        _table("test_questions")
        .select("id", count="exact")
        .eq("school_id", school_id)
        .eq("section_id", section_id)
        .is_("deleted_at", "null")
        .limit(0)
        .execute()
    )
    (
        _table("test_sections")
        .update({"question_count": int(response.count or 0)})
        .eq("school_id", school_id)
        .eq("id", section_id)
        .execute()
    )


def _question_rows_for_test(school_id: str, test_id: str) -> list[dict[str, Any]]:
    rows = list(
        _table("test_questions")
        .select("*")
        .eq("school_id", school_id)
        .eq("test_id", test_id)
        .is_("deleted_at", "null")
        .order("section_id", desc=False)
        .order("display_order", desc=False)
        .execute()
        .data
        or []
    )
    return [dict(row) for row in rows]


def _student_rows_by_ids(school_id: str, student_ids: list[str]) -> dict[str, dict[str, Any]]:
    cleaned_ids = [_normalize(student_id) for student_id in student_ids if _normalize(student_id)]
    if not cleaned_ids:
        return {}
    rows = list(
        _client()
        .table("students")
        .select("id, batch_id, full_name")
        .eq("school_id", school_id)
        .in_("id", cleaned_ids)
        .execute()
        .data
        or []
    )
    return {_normalize(row.get("id")): dict(row) for row in rows}


def _attempt_responses_rows(school_id: str, attempt_id: str) -> list[dict[str, Any]]:
    rows = list(
        _table("test_responses")
        .select("*")
        .eq("school_id", school_id)
        .eq("attempt_id", attempt_id)
        .is_("deleted_at", "null")
        .order("created_at", desc=False)
        .execute()
        .data
        or []
    )
    return [dict(row) for row in rows]


def _extract_candidate_answers(payload: dict[str, Any]) -> list[str]:
    if not isinstance(payload, dict):
        return []
    candidates: list[str] = []
    for key in ("selected_option_id", "selected_option", "value", "answer", "text"):
        value = payload.get(key)
        if value not in (None, ""):
            candidates.append(_normalize(value).lower())
    values = payload.get("selected_option_ids")
    if isinstance(values, list):
        candidates.extend(_normalize(item).lower() for item in values if _normalize(item))
    return [item for item in candidates if item]


def _extract_expected_answers(answer_key: dict[str, Any]) -> list[str]:
    if not isinstance(answer_key, dict):
        return []
    candidates: list[str] = []
    for key in ("correct_option_id", "correct_value", "expected_value"):
        value = answer_key.get(key)
        if value not in (None, ""):
            candidates.append(_normalize(value).lower())
    values = answer_key.get("correct_option_ids") or answer_key.get("accepted_values")
    if isinstance(values, list):
        candidates.extend(_normalize(item).lower() for item in values if _normalize(item))
    if "expected_bool" in answer_key:
        candidates.append(str(bool(answer_key.get("expected_bool"))).lower())
    return [item for item in candidates if item]


def _score_response(question: dict[str, Any], response_payload: dict[str, Any]) -> tuple[bool | None, float]:
    question_type = _normalize(question.get("question_type"))
    marks = float(question.get("marks") or 0)
    negative_marks = float(question.get("negative_marks") or 0)
    if question_type in {"short_answer", "long_answer"}:
        return None, 0.0
    given = sorted(set(_extract_candidate_answers(response_payload)))
    expected = sorted(set(_extract_expected_answers(_normalize_json_object(question.get("answer_key")))))
    if not given:
        return False, 0.0
    is_correct = given == expected and bool(expected)
    return is_correct, marks if is_correct else max(0.0, -negative_marks)


def _recalculate_result_ranks(school_id: str, test_id: str) -> None:
    result_rows = [
        dict(row)
        for row in list(
            _table("test_results")
            .select("id, student_id, score_obtained, percentage, published_at, created_at")
            .eq("school_id", school_id)
            .eq("test_id", test_id)
            .is_("deleted_at", "null")
            .execute()
            .data
            or []
        )
    ]
    if not result_rows:
        return

    student_map = _student_rows_by_ids(school_id, [_normalize(row.get("student_id")) for row in result_rows])

    def _sort_key(row: dict[str, Any]) -> tuple[float, float, str, str]:
        return (
            float(row.get("score_obtained") or 0),
            float(row.get("percentage") or 0),
            _normalize(row.get("published_at") or row.get("created_at")),
            _normalize(row.get("id")),
        )

    ranked_school = sorted(result_rows, key=_sort_key, reverse=True)
    for school_rank, row in enumerate(ranked_school, start=1):
        _table("test_results").update({"rank_in_school": school_rank}).eq("school_id", school_id).eq(
            "id", _normalize(row.get("id"))
        ).execute()

    batch_groups: dict[str, list[dict[str, Any]]] = {}
    for row in result_rows:
        batch_id = _normalize(student_map.get(_normalize(row.get("student_id")), {}).get("batch_id")) or "unassigned"
        batch_groups.setdefault(batch_id, []).append(row)

    for batch_rows in batch_groups.values():
        for batch_rank, row in enumerate(sorted(batch_rows, key=_sort_key, reverse=True), start=1):
            _table("test_results").update({"rank_in_batch": batch_rank}).eq("school_id", school_id).eq(
                "id", _normalize(row.get("id"))
            ).execute()


def _build_duplicate_test_code(school_id: str, base_code: str | None) -> str | None:
    normalized_base = _normalize(base_code)
    if not normalized_base:
        return None
    suffix_seed = datetime.now(timezone.utc).strftime("%H%M%S")
    candidate = f"{normalized_base}-COPY-{suffix_seed}"
    existing = list(
        _table("tests")
        .select("id")
        .eq("school_id", school_id)
        .eq("test_code", candidate)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    return candidate if not existing else f"{candidate}-{datetime.now(timezone.utc).strftime('%f')[:3]}"


def _assert_student_can_access_test(student: dict[str, Any], test: dict[str, Any]) -> None:
    status = _normalize(test.get("status"))
    if not bool(test.get("is_active", True)) or status not in {"published", "in_progress", "completed", "closed"}:
        raise HTTPException(status_code=403, detail="This test is not available for student access")
    test_batch_id = _normalize(test.get("batch_id"))
    student_batch_id = _normalize(student.get("batch_id"))
    if test_batch_id and student_batch_id and test_batch_id != student_batch_id:
        raise HTTPException(status_code=403, detail="This test is not assigned to the student's batch")


def _normalize_option_items(raw_options: list[dict[str, Any]] | list[str]) -> list[dict[str, Any]]:
    option_items: list[dict[str, Any]] = []
    for index, item in enumerate(raw_options, start=1):
        if isinstance(item, dict):
            label = _normalize(item.get("label") or item.get("text") or item.get("value") or item.get("id"))
            option_id = _normalize(item.get("id")) or f"option_{index}"
            option_payload = {"id": option_id, "label": label or option_id, "value": label or option_id}
            image_url = _normalize(item.get("image_url"))
            if image_url:
                option_payload["image_url"] = image_url
            option_items.append(option_payload)
        else:
            label = _normalize(item)
            if label:
                option_items.append({"id": f"option_{index}", "label": label, "value": label})
    return option_items


def _question_bank_row_to_question_payload(row: dict[str, Any], test_id: str, section_id: str, display_order: int) -> dict[str, Any]:
    return {
        "test_id": test_id,
        "section_id": section_id,
        "display_order": display_order,
        "question_type": _normalize(row.get("question_type")) or "single_choice",
        "difficulty_level": _normalize(row.get("difficulty_level")) or "medium",
        "prompt_text": _normalize(row.get("prompt_text")),
        "option_items": _normalize_option_items(list(row.get("option_items") or [])),
        "answer_key": _normalize_json_object(row.get("answer_key")),
        "explanation": row.get("explanation"),
        "marks": float(row.get("marks") or 1),
        "negative_marks": float(row.get("negative_marks") or 0),
        "metadata": _normalize_json_object(row.get("metadata")),
    }


def _question_bank_import_row_to_payload(row: dict[str, Any]) -> dict[str, Any]:
    question = _normalize(row.get("Question"))
    if not question:
        raise HTTPException(status_code=400, detail="Question column is required for every imported row")
    options = []
    for key in ("Option A", "Option B", "Option C", "Option D"):
        value = _normalize(row.get(key))
        if value:
            options.append(value)
    correct_answer = _normalize(row.get("Correct Answer"))
    difficulty = _normalize(row.get("Difficulty")).lower() or "medium"
    prompt_image_url = _normalize(row.get("Question Image URL"))
    option_image_map = {
        "option_1": _normalize(row.get("Option A Image URL")),
        "option_2": _normalize(row.get("Option B Image URL")),
        "option_3": _normalize(row.get("Option C Image URL")),
        "option_4": _normalize(row.get("Option D Image URL")),
    }
    option_items = _normalize_option_items(options)
    for option in option_items:
        image_url = option_image_map.get(_normalize(option.get("id")))
        if image_url:
            option["image_url"] = image_url
    matched = next(
        (
            item["id"]
            for item in option_items
            if _normalize(item.get("label")).lower() == correct_answer.lower()
            or _normalize(item.get("id")).lower() == correct_answer.lower()
        ),
        None,
    )
    if option_items:
        answer_key = {"correct_option_id": matched or correct_answer}
        question_type = "single_choice"
    else:
        answer_key = {"accepted_values": [correct_answer]} if correct_answer else {}
        question_type = "short_answer"
    return {
        "subject": _normalize(row.get("Subject")) or None,
        "chapter": _normalize(row.get("Chapter")) or None,
        "topic": _normalize(row.get("Topic")) or None,
        "question_type": question_type,
        "difficulty_level": difficulty if difficulty in {"easy", "medium", "hard"} else "medium",
        "prompt_text": question,
        "option_items": option_items,
        "answer_key": answer_key,
        "explanation": _normalize(row.get("Explanation")) or None,
        "marks": float(row.get("Marks") or 1),
        "negative_marks": float(row.get("Negative Marks") or 0),
        "metadata": {
            "source": "excel_import",
            "question_image_url": prompt_image_url or None,
        },
    }


def list_question_bank(
    school_id: str,
    *,
    subject: str | None = None,
    chapter: str | None = None,
    topic: str | None = None,
    difficulty_level: str | None = None,
    skip: int = 0,
    limit: int = 200,
) -> list[dict[str, Any]]:
    query = _question_bank_table().select("*").eq("school_id", school_id).is_("deleted_at", "null").eq("is_active", True)
    if subject:
        query = query.eq("subject", subject)
    if chapter:
        query = query.eq("chapter", chapter)
    if topic:
        query = query.eq("topic", topic)
    if difficulty_level:
        query = query.eq("difficulty_level", difficulty_level)
    rows = list(query.order("created_at", desc=True).range(max(skip, 0), max(skip, 0) + max(limit, 1) - 1).execute().data or [])
    return [_serialize_question_bank_item(dict(row)) for row in rows]


def create_question_bank_item(school_id: str, profile_id: str | None, payload: dict[str, Any]) -> dict[str, Any]:
    prompt_text = _normalize(payload.get("prompt_text"))
    if not prompt_text:
        raise HTTPException(status_code=400, detail="prompt_text is required")
    response = _question_bank_table().insert(
        {
            "school_id": school_id,
            "created_by_profile_id": _normalize_optional_uuid(profile_id),
            "updated_by_profile_id": _normalize_optional_uuid(profile_id),
            "subject": _normalize(payload.get("subject")) or None,
            "chapter": _normalize(payload.get("chapter")) or None,
            "topic": _normalize(payload.get("topic")) or None,
            "question_type": _normalize(payload.get("question_type")) or "single_choice",
            "difficulty_level": _normalize(payload.get("difficulty_level")) or "medium",
            "prompt_text": prompt_text,
            "option_items": _normalize_option_items(list(payload.get("option_items") or [])),
            "answer_key": _normalize_json_object(payload.get("answer_key")),
            "explanation": payload.get("explanation"),
            "marks": float(payload.get("marks") or 1),
            "negative_marks": float(payload.get("negative_marks") or 0),
            "metadata": _normalize_json_object(payload.get("metadata")),
            "is_active": True,
        }
    ).execute()
    rows = list(response.data or [])
    if not rows:
        raise HTTPException(status_code=500, detail="Question bank create returned no row")
    return _serialize_question_bank_item(dict(rows[0]))


def import_question_bank_workbook(school_id: str, profile_id: str | None, file_bytes: bytes) -> dict[str, Any]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    required_headers = {"Question", "Correct Answer", "Difficulty", "Topic", "Chapter"}
    missing = sorted(required_headers - set(headers))
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    created_items: list[dict[str, Any]] = []
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue
        row = {headers[index]: values[index] for index in range(len(headers))}
        try:
            created_items.append(create_question_bank_item(school_id, profile_id, _question_bank_import_row_to_payload(row)))
        except HTTPException as exc:
            raise HTTPException(status_code=400, detail=f"Row {row_index}: {exc.detail}") from exc
    return {"created_count": len(created_items), "items": created_items}


def generate_ai_test(school_id: str, profile_id: str | None, payload: dict[str, Any]) -> dict[str, Any]:
    question_count = max(1, min(int(payload.get("question_count") or 10), 50))
    generated = generate_json(
        "Generate a valid JSON object for an online coaching test with keys title, description, instructions, questions. "
        "Each question must include prompt_text, question_type, difficulty_level, option_items, answer_key, explanation, marks, negative_marks, metadata. "
        f"Subject: {_normalize(payload.get('subject'))}. "
        f"Chapter: {_normalize(payload.get('chapter'))}. "
        f"Topic: {_normalize(payload.get('topic'))}. "
        f"Difficulty: {_normalize(payload.get('difficulty')) or 'medium'}. "
        f"Question count: {question_count}. "
        "Use single_choice by default. Return only JSON."
    )
    questions = list(generated.get("questions") or [])
    if not questions:
        raise AIProviderError("AI service temporarily unavailable")

    created_test = create_test(
        school_id,
        profile_id,
        {
            "title": _normalize(payload.get("title")) or _normalize(generated.get("title")) or f"{_normalize(payload.get('topic'))} AI Test",
            "description": generated.get("description") or f"AI generated test for {_normalize(payload.get('subject'))}",
            "instructions": generated.get("instructions") or "Attempt all questions carefully.",
            "batch_id": _normalize(payload.get("batch_id")) or None,
            "status": "draft",
            "duration_minutes": int(payload.get("duration_minutes") or 60),
            "total_marks": round(sum(float(item.get("marks") or payload.get("marks_per_question") or 1) for item in questions), 2),
            "metadata": {
                "source": "ai_generator",
                "subject": _normalize(payload.get("subject")),
                "chapter": _normalize(payload.get("chapter")),
                "topic": _normalize(payload.get("topic")),
                "difficulty": _normalize(payload.get("difficulty")) or "medium",
            },
        },
    )
    section_id = _normalize((created_test.get("sections") or [{}])[0].get("id"))
    created_questions: list[dict[str, Any]] = []
    for index, item in enumerate(questions, start=1):
        prompt_text = _normalize(item.get("prompt_text"))
        if not prompt_text:
            continue
        created_questions.append(
            create_question(
                school_id,
                {
                    "test_id": created_test.get("id"),
                    "section_id": section_id,
                    "display_order": index,
                    "question_type": _normalize(item.get("question_type")) or "single_choice",
                    "difficulty_level": _normalize(item.get("difficulty_level")) or _normalize(payload.get("difficulty")) or "medium",
                    "prompt_text": prompt_text,
                    "option_items": _normalize_option_items(list(item.get("option_items") or [])),
                    "answer_key": _normalize_json_object(item.get("answer_key")),
                    "explanation": item.get("explanation"),
                    "marks": float(item.get("marks") or payload.get("marks_per_question") or 1),
                    "negative_marks": float(item.get("negative_marks") or payload.get("negative_marks") or 0),
                    "metadata": _normalize_json_object(item.get("metadata")),
                },
                profile_id,
            )
        )
    return {"success": True, "test": get_test(school_id, _normalize(created_test.get("id"))), "questions": created_questions}


def list_tests(
    school_id: str,
    *,
    include_inactive: bool = False,
    student_batch_id: str | None = None,
    skip: int = 0,
    limit: int = 100,
) -> list[dict[str, Any]]:
    query = _table("tests").select("*").eq("school_id", school_id).is_("deleted_at", "null")
    if not include_inactive:
        query = query.eq("is_active", True)
    if student_batch_id:
        query = query.or_(f"batch_id.eq.{student_batch_id},batch_id.is.null")
        query = query.in_("status", ["published", "in_progress", "completed", "closed"])
    rows = list(
        query
        .order("created_at", desc=True)
        .range(max(skip, 0), max(skip, 0) + max(limit, 1) - 1)
        .execute()
        .data
        or []
    )
    sections_map = _load_sections_map(school_id, [_normalize(row.get("id")) for row in rows])
    return [_serialize_test(dict(row), sections_map.get(_normalize(row.get("id")), [])) for row in rows]


def get_test(school_id: str, test_id: str) -> dict[str, Any]:
    row = _get_test_row(school_id, test_id)
    sections = _load_sections_map(school_id, [test_id]).get(test_id, [])
    return _serialize_test(row, sections)


def get_test_for_student(school_id: str, test_id: str, profile_id: str) -> dict[str, Any]:
    student = _get_student_by_profile_id(school_id, profile_id)
    test = _get_test_row(school_id, test_id)
    _assert_student_can_access_test(student, test)
    sections = _load_sections_map(school_id, [test_id]).get(test_id, [])
    return _serialize_test(test, sections)


def create_test(school_id: str, profile_id: str | None, payload: dict[str, Any]) -> dict[str, Any]:
    title = _normalize(payload.get("title"))
    if not title:
        raise HTTPException(status_code=400, detail="Test title is required")
    desired_status = _normalize(payload.get("status")) or "draft"
    insert_payload = {
        "school_id": school_id,
        "subject_id": _normalize_optional_uuid(payload.get("subject_id")),
        "batch_id": _normalize_optional_uuid(payload.get("batch_id")),
        "created_by_profile_id": _normalize_optional_uuid(profile_id),
        "published_by_profile_id": _normalize_optional_uuid(profile_id) if desired_status == "published" else None,
        "test_code": _normalize(payload.get("test_code")) or None,
        "title": title,
        "description": payload.get("description"),
        "instructions": payload.get("instructions"),
        "test_type": _normalize(payload.get("test_type")) or "objective",
        "delivery_mode": _normalize(payload.get("delivery_mode")) or "scheduled",
        "status": desired_status,
        "duration_minutes": int(payload.get("duration_minutes") or 60),
        "total_marks": float(payload.get("total_marks") or 0),
        "pass_marks": float(payload.get("pass_marks")) if payload.get("pass_marks") is not None else None,
        "max_attempts": int(payload.get("max_attempts") or 1),
        "shuffle_questions": bool(payload.get("shuffle_questions", False)),
        "shuffle_options": bool(payload.get("shuffle_options", False)),
        "show_result_immediately": bool(payload.get("show_result_immediately", False)),
        "allow_review": bool(payload.get("allow_review", True)),
        "starts_at": _to_iso_datetime(payload.get("starts_at")),
        "ends_at": _to_iso_datetime(payload.get("ends_at")),
        "published_at": _utc_now_iso() if desired_status == "published" else None,
        "metadata": _normalize_json_object(payload.get("metadata")),
        "is_active": bool(payload.get("is_active", True)),
    }
    response = _table("tests").insert(insert_payload).execute()
    rows = list(response.data or [])
    if not rows:
        raise HTTPException(status_code=500, detail="Test create returned no row")
    created = dict(rows[0])
    _ensure_default_section(school_id, _normalize(created.get("id")))
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.test_created",
        entity_id=_normalize(created.get("id")),
        payload={"title": title, "status": desired_status},
    )
    return get_test(school_id, _normalize(created.get("id")))


def update_test(school_id: str, test_id: str, profile_id: str | None, payload: dict[str, Any]) -> dict[str, Any]:
    existing = _get_test_row(school_id, test_id)
    update_payload: dict[str, Any] = {}
    for key in (
        "title",
        "description",
        "instructions",
        "test_code",
        "test_type",
        "delivery_mode",
        "status",
        "shuffle_questions",
        "shuffle_options",
        "show_result_immediately",
        "allow_review",
        "is_active",
    ):
        if key in payload:
            update_payload[key] = payload.get(key)
    if "subject_id" in payload:
        update_payload["subject_id"] = _normalize_optional_uuid(payload.get("subject_id"))
    if "batch_id" in payload:
        update_payload["batch_id"] = _normalize_optional_uuid(payload.get("batch_id"))
    if "duration_minutes" in payload and payload.get("duration_minutes") is not None:
        update_payload["duration_minutes"] = int(payload.get("duration_minutes"))
    if "total_marks" in payload and payload.get("total_marks") is not None:
        update_payload["total_marks"] = float(payload.get("total_marks"))
    if "pass_marks" in payload:
        update_payload["pass_marks"] = float(payload.get("pass_marks")) if payload.get("pass_marks") is not None else None
    if "max_attempts" in payload and payload.get("max_attempts") is not None:
        update_payload["max_attempts"] = int(payload.get("max_attempts"))
    if "starts_at" in payload:
        update_payload["starts_at"] = _to_iso_datetime(payload.get("starts_at"))
    if "ends_at" in payload:
        update_payload["ends_at"] = _to_iso_datetime(payload.get("ends_at"))
    if "metadata" in payload and payload.get("metadata") is not None:
        update_payload["metadata"] = _normalize_json_object(payload.get("metadata"))

    new_status = _normalize(update_payload.get("status") or existing.get("status"))
    old_status = _normalize(existing.get("status"))
    if new_status == "published" and old_status != "published":
        update_payload["published_at"] = _utc_now_iso()
        update_payload["published_by_profile_id"] = _normalize_optional_uuid(profile_id)
    elif old_status == "published" and new_status != "published":
        update_payload["published_at"] = None

    _table("tests").update(update_payload).eq("school_id", school_id).eq("id", test_id).execute()
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.test_updated",
        entity_id=test_id,
        payload={"updated_fields": sorted(update_payload.keys())},
    )
    return get_test(school_id, test_id)


def publish_test(school_id: str, test_id: str, profile_id: str | None) -> dict[str, Any]:
    _get_test_row(school_id, test_id)
    update_payload = {
        "status": "published",
        "published_at": _utc_now_iso(),
        "published_by_profile_id": _normalize_optional_uuid(profile_id),
        "is_active": True,
    }
    _table("tests").update(update_payload).eq("school_id", school_id).eq("id", test_id).execute()
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.test_published",
        entity_id=test_id,
        payload={"status": "published"},
    )
    return get_test(school_id, test_id)


def unpublish_test(school_id: str, test_id: str, profile_id: str | None) -> dict[str, Any]:
    _get_test_row(school_id, test_id)
    _table("tests").update({"status": "draft", "published_at": None}).eq("school_id", school_id).eq("id", test_id).execute()
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.test_unpublished",
        entity_id=test_id,
        payload={"status": "draft"},
    )
    return get_test(school_id, test_id)


def close_test(school_id: str, test_id: str, profile_id: str | None) -> dict[str, Any]:
    _get_test_row(school_id, test_id)
    _table("tests").update({"status": "closed"}).eq("school_id", school_id).eq("id", test_id).execute()
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.test_closed",
        entity_id=test_id,
        payload={"status": "closed"},
    )
    return get_test(school_id, test_id)


def duplicate_test(school_id: str, test_id: str, profile_id: str | None) -> dict[str, Any]:
    source_test = _get_test_row(school_id, test_id)
    source_sections = list(
        _table("test_sections")
        .select("*")
        .eq("school_id", school_id)
        .eq("test_id", test_id)
        .is_("deleted_at", "null")
        .order("display_order", desc=False)
        .execute()
        .data
        or []
    )
    source_questions = _question_rows_for_test(school_id, test_id)

    duplicated_test = create_test(
        school_id,
        profile_id,
        {
            "title": f"{_normalize(source_test.get('title'))} (Copy)",
            "description": source_test.get("description"),
            "instructions": source_test.get("instructions"),
            "test_code": _build_duplicate_test_code(school_id, source_test.get("test_code")),
            "subject_id": source_test.get("subject_id"),
            "batch_id": source_test.get("batch_id"),
            "test_type": source_test.get("test_type"),
            "delivery_mode": source_test.get("delivery_mode"),
            "status": "draft",
            "duration_minutes": source_test.get("duration_minutes"),
            "total_marks": source_test.get("total_marks"),
            "pass_marks": source_test.get("pass_marks"),
            "max_attempts": source_test.get("max_attempts"),
            "shuffle_questions": source_test.get("shuffle_questions"),
            "shuffle_options": source_test.get("shuffle_options"),
            "show_result_immediately": source_test.get("show_result_immediately"),
            "allow_review": source_test.get("allow_review"),
            "starts_at": source_test.get("starts_at"),
            "ends_at": source_test.get("ends_at"),
            "metadata": _normalize_json_object(source_test.get("metadata")),
        },
    )

    duplicated_test_id = _normalize(duplicated_test.get("id"))
    default_section_id = _normalize((duplicated_test.get("sections") or [{}])[0].get("id"))
    section_map: dict[str, str] = {}
    ordered_source_sections = [dict(row) for row in source_sections]
    if ordered_source_sections:
        first_section = ordered_source_sections[0]
        _table("test_sections").update(
            {
                "section_code": first_section.get("section_code"),
                "title": first_section.get("title"),
                "description": first_section.get("description"),
                "display_order": first_section.get("display_order"),
                "question_type": first_section.get("question_type"),
                "marks_per_question": first_section.get("marks_per_question"),
                "negative_marks": first_section.get("negative_marks"),
                "question_count": 0,
            }
        ).eq("school_id", school_id).eq("id", default_section_id).execute()
        section_map[_normalize(first_section.get("id"))] = default_section_id

        for section in ordered_source_sections[1:]:
            response = _table("test_sections").insert(
                {
                    "school_id": school_id,
                    "test_id": duplicated_test_id,
                    "section_code": section.get("section_code"),
                    "title": section.get("title"),
                    "description": section.get("description"),
                    "display_order": section.get("display_order"),
                    "question_type": section.get("question_type"),
                    "marks_per_question": section.get("marks_per_question"),
                    "negative_marks": section.get("negative_marks"),
                    "question_count": 0,
                    "is_active": bool(section.get("is_active", True)),
                }
            ).execute()
            rows = list(response.data or [])
            if not rows:
                raise HTTPException(status_code=500, detail="Failed to duplicate test section")
            section_map[_normalize(section.get("id"))] = _normalize(rows[0].get("id"))

    for question in source_questions:
        source_section_id = _normalize(question.get("section_id"))
        target_section_id = section_map.get(source_section_id) or default_section_id
        _table("test_questions").insert(
            {
                "school_id": school_id,
                "test_id": duplicated_test_id,
                "section_id": target_section_id,
                "question_code": question.get("question_code"),
                "display_order": question.get("display_order"),
                "question_type": question.get("question_type"),
                "difficulty_level": question.get("difficulty_level"),
                "prompt_text": question.get("prompt_text"),
                "option_items": list(question.get("option_items") or []),
                "answer_key": _normalize_json_object(question.get("answer_key")),
                "explanation": question.get("explanation"),
                "marks": question.get("marks"),
                "negative_marks": question.get("negative_marks"),
                "metadata": _normalize_json_object(question.get("metadata")),
                "is_active": bool(question.get("is_active", True)),
            }
        ).execute()
        _recount_section_questions(school_id, target_section_id)

    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.test_duplicated",
        entity_id=duplicated_test_id,
        payload={"source_test_id": test_id},
    )
    return get_test(school_id, duplicated_test_id)


def delete_test(school_id: str, test_id: str, profile_id: str | None) -> dict[str, Any]:
    _get_test_row(school_id, test_id)
    now = _utc_now_iso()
    deleted_by = _normalize_optional_uuid(profile_id)
    _table("tests").update(
        {
            "is_active": False,
            "deleted_at": now,
            "deleted_by_profile_id": deleted_by,
            "status": "archived",
        }
    ).eq("school_id", school_id).eq("id", test_id).execute()
    _table("test_sections").update(
        {"is_active": False, "deleted_at": now, "deleted_by_profile_id": deleted_by}
    ).eq("school_id", school_id).eq("test_id", test_id).is_("deleted_at", "null").execute()
    _table("test_questions").update(
        {"is_active": False, "deleted_at": now, "deleted_by_profile_id": deleted_by}
    ).eq("school_id", school_id).eq("test_id", test_id).is_("deleted_at", "null").execute()
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.test_deleted",
        entity_id=test_id,
        payload={"soft_delete": True},
    )
    return {"message": "Online test deleted successfully"}


def list_questions(
    school_id: str,
    *,
    test_id: str,
    section_id: str | None = None,
    skip: int = 0,
    limit: int = 100,
) -> list[dict[str, Any]]:
    _get_test_row(school_id, test_id)
    query = (
        _table("test_questions")
        .select("*")
        .eq("school_id", school_id)
        .eq("test_id", test_id)
        .is_("deleted_at", "null")
    )
    if section_id:
        query = query.eq("section_id", section_id)
    rows = list(
        query
        .order("display_order", desc=False)
        .range(max(skip, 0), max(skip, 0) + max(limit, 1) - 1)
        .execute()
        .data
        or []
    )
    return [_serialize_question(dict(row)) for row in rows]


def list_questions_for_student(
    school_id: str,
    *,
    test_id: str,
    profile_id: str,
    section_id: str | None = None,
    skip: int = 0,
    limit: int = 100,
) -> list[dict[str, Any]]:
    student = _get_student_by_profile_id(school_id, profile_id)
    test = _get_test_row(school_id, test_id)
    _assert_student_can_access_test(student, test)
    questions = list_questions(school_id, test_id=test_id, section_id=section_id, skip=skip, limit=limit)
    return [_sanitize_question_for_student(question) for question in questions]


def create_question(school_id: str, payload: dict[str, Any], profile_id: str | None = None) -> dict[str, Any]:
    test_id = _normalize(payload.get("test_id"))
    if not test_id:
        raise HTTPException(status_code=400, detail="test_id is required")
    _get_test_row(school_id, test_id)
    section_id = _resolve_section_id(school_id, test_id, _normalize(payload.get("section_id")) or None)
    prompt_text = _normalize(payload.get("prompt_text"))
    if not prompt_text:
        raise HTTPException(status_code=400, detail="prompt_text is required")
    insert_payload = {
        "school_id": school_id,
        "test_id": test_id,
        "section_id": section_id,
        "question_code": _normalize(payload.get("question_code")) or None,
        "display_order": int(payload.get("display_order") or 1),
        "question_type": _normalize(payload.get("question_type")) or "single_choice",
        "difficulty_level": _normalize(payload.get("difficulty_level")) or "medium",
        "prompt_text": prompt_text,
        "option_items": list(payload.get("option_items") or []),
        "answer_key": _normalize_json_object(payload.get("answer_key")),
        "explanation": payload.get("explanation"),
        "marks": float(payload.get("marks") or 1),
        "negative_marks": float(payload.get("negative_marks") or 0),
        "metadata": _normalize_json_object(payload.get("metadata")),
        "is_active": bool(payload.get("is_active", True)),
    }
    response = _table("test_questions").insert(insert_payload).execute()
    rows = list(response.data or [])
    if not rows:
        raise HTTPException(status_code=500, detail="Question create returned no row")
    created = dict(rows[0])
    _recount_section_questions(school_id, section_id)
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.question_created",
        entity_id=_normalize(created.get("id")),
        payload={"test_id": test_id, "section_id": section_id},
    )
    return _serialize_question(created)


def update_question(school_id: str, question_id: str, payload: dict[str, Any], profile_id: str | None = None) -> dict[str, Any]:
    existing_rows = list(
        _table("test_questions")
        .select("*")
        .eq("school_id", school_id)
        .eq("id", question_id)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    if not existing_rows:
        raise HTTPException(status_code=404, detail="Question not found")
    existing = dict(existing_rows[0])
    update_payload: dict[str, Any] = {}
    for key in ("question_code", "prompt_text", "explanation", "question_type", "difficulty_level", "is_active"):
        if key in payload:
            update_payload[key] = payload.get(key)
    if "display_order" in payload and payload.get("display_order") is not None:
        update_payload["display_order"] = int(payload.get("display_order"))
    if "option_items" in payload and payload.get("option_items") is not None:
        update_payload["option_items"] = list(payload.get("option_items") or [])
    if "answer_key" in payload and payload.get("answer_key") is not None:
        update_payload["answer_key"] = _normalize_json_object(payload.get("answer_key"))
    if "marks" in payload and payload.get("marks") is not None:
        update_payload["marks"] = float(payload.get("marks"))
    if "negative_marks" in payload and payload.get("negative_marks") is not None:
        update_payload["negative_marks"] = float(payload.get("negative_marks"))
    if "metadata" in payload and payload.get("metadata") is not None:
        update_payload["metadata"] = _normalize_json_object(payload.get("metadata"))
    if "section_id" in payload:
        update_payload["section_id"] = _resolve_section_id(
            school_id,
            _normalize(existing.get("test_id")),
            payload.get("section_id"),
        )
    _table("test_questions").update(update_payload).eq("school_id", school_id).eq("id", question_id).execute()
    if update_payload.get("section_id") and _normalize(update_payload.get("section_id")) != _normalize(existing.get("section_id")):
        _recount_section_questions(school_id, _normalize(existing.get("section_id")))
        _recount_section_questions(school_id, _normalize(update_payload.get("section_id")))
    else:
        _recount_section_questions(school_id, _normalize(existing.get("section_id")))
    refreshed = list(
        _table("test_questions")
        .select("*")
        .eq("school_id", school_id)
        .eq("id", question_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.question_updated",
        entity_id=question_id,
        payload={"updated_fields": sorted(update_payload.keys())},
    )
    return _serialize_question(dict(refreshed[0]))


def delete_question(school_id: str, question_id: str, profile_id: str | None) -> dict[str, Any]:
    existing_rows = list(
        _table("test_questions")
        .select("id, section_id, test_id")
        .eq("school_id", school_id)
        .eq("id", question_id)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    if not existing_rows:
        raise HTTPException(status_code=404, detail="Question not found")
    section_id = _normalize(existing_rows[0].get("section_id"))
    _table("test_questions").update(
        {
            "is_active": False,
            "deleted_at": _utc_now_iso(),
            "deleted_by_profile_id": _normalize_optional_uuid(profile_id),
        }
    ).eq("school_id", school_id).eq("id", question_id).execute()
    _recount_section_questions(school_id, section_id)
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.question_deleted",
        entity_id=question_id,
        payload={"test_id": _normalize(existing_rows[0].get("test_id")), "soft_delete": True},
    )
    return {"message": "Question deleted successfully"}


def list_attempts(
    school_id: str,
    *,
    student_id: str | None = None,
    test_id: str | None = None,
    skip: int = 0,
    limit: int = 100,
) -> list[dict[str, Any]]:
    query = _table("test_attempts").select("*").eq("school_id", school_id).is_("deleted_at", "null")
    if student_id:
        query = query.eq("student_id", student_id)
    if test_id:
        query = query.eq("test_id", test_id)
    rows = list(
        query
        .order("created_at", desc=True)
        .range(max(skip, 0), max(skip, 0) + max(limit, 1) - 1)
        .execute()
        .data
        or []
    )
    return [_serialize_attempt(dict(row)) for row in rows]


def get_attempt(school_id: str, attempt_id: str) -> dict[str, Any]:
    row = _get_attempt_row(school_id, attempt_id)
    responses = [_serialize_response(item) for item in _attempt_responses_rows(school_id, attempt_id)]
    return _serialize_attempt(row, responses)


def start_attempt(school_id: str, test_id: str, profile_id: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    next_payload = {"test_id": test_id}
    if isinstance(payload, dict):
        next_payload.update(payload)
    return create_attempt(school_id, profile_id, next_payload)


def create_attempt(school_id: str, profile_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    from app.services.supabase_account_security import start_test_session

    student = _get_student_by_profile_id(school_id, profile_id)
    test_id = _normalize(payload.get("test_id"))
    if not test_id:
        raise HTTPException(status_code=400, detail="test_id is required")
    test = _get_test_row(school_id, test_id)
    _assert_student_can_access_test(student, test)

    existing_in_progress = list(
        _table("test_attempts")
        .select("id")
        .eq("school_id", school_id)
        .eq("test_id", test_id)
        .eq("student_id", _normalize(student.get("id")))
        .eq("status", "in_progress")
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    if existing_in_progress:
        start_test_session(
            school_id=school_id,
            test_id=test_id,
            attempt_id=_normalize(existing_in_progress[0].get("id")),
            student_id=_normalize(student.get("id")),
            profile_id=profile_id,
            session_key=_normalize(payload.get("session_key")) or None,
            device_id=_normalize(payload.get("device_id")) or None,
            mode=_normalize(payload.get("session_mode")) or "terminate_previous",
        )
        return get_attempt(school_id, _normalize(existing_in_progress[0].get("id")))

    existing_attempts = list(
        _table("test_attempts")
        .select("id")
        .eq("school_id", school_id)
        .eq("test_id", test_id)
        .eq("student_id", _normalize(student.get("id")))
        .is_("deleted_at", "null")
        .execute()
        .data
        or []
    )
    max_attempts = int(test.get("max_attempts") or 1)
    if len(existing_attempts) >= max_attempts:
        raise HTTPException(status_code=400, detail="Maximum attempts reached for this test")
    total_questions = len(_question_rows_for_test(school_id, test_id))
    attempt_number = len(existing_attempts) + 1
    response = _table("test_attempts").insert(
        {
            "school_id": school_id,
            "test_id": test_id,
            "student_id": _normalize(student.get("id")),
            "attempt_number": attempt_number,
            "status": "in_progress",
            "started_at": _utc_now_iso(),
            "total_questions_snapshot": total_questions,
            "answered_questions_snapshot": 0,
            "time_spent_seconds": 0,
            "metadata": {"student_name": student.get("full_name")},
            "is_active": True,
        }
    ).execute()
    rows = list(response.data or [])
    if not rows:
        raise HTTPException(status_code=500, detail="Attempt start returned no row")
    created = dict(rows[0])
    start_test_session(
        school_id=school_id,
        test_id=test_id,
        attempt_id=_normalize(created.get("id")),
        student_id=_normalize(student.get("id")),
        profile_id=profile_id,
        session_key=_normalize(payload.get("session_key")) or None,
        device_id=_normalize(payload.get("device_id")) or None,
        mode=_normalize(payload.get("session_mode")) or "terminate_previous",
    )
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.attempt_started",
        entity_id=_normalize(created.get("id")),
        payload={"test_id": test_id, "attempt_number": attempt_number},
    )
    return _serialize_attempt(created, [])


def save_attempt(school_id: str, attempt_id: str, profile_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    from app.services.supabase_account_security import touch_test_session

    attempt = _get_attempt_row(school_id, attempt_id)
    student = _get_student_by_profile_id(school_id, profile_id)
    if _normalize(attempt.get("student_id")) != _normalize(student.get("id")):
        raise HTTPException(status_code=403, detail="Students can update only their own attempts")
    if _normalize(attempt.get("status")) != "in_progress":
        raise HTTPException(status_code=400, detail="Only in-progress attempts can receive responses")
    question_id = _normalize(payload.get("question_id"))
    if not question_id:
        raise HTTPException(status_code=400, detail="question_id is required")
    question_rows = list(
        _table("test_questions")
        .select("*")
        .eq("school_id", school_id)
        .eq("id", question_id)
        .eq("test_id", _normalize(attempt.get("test_id")))
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    if not question_rows:
        raise HTTPException(status_code=404, detail="Question not found for this attempt")
    response_payload = _normalize_json_object(payload.get("response_payload"))
    existing_rows = list(
        _table("test_responses")
        .select("*")
        .eq("school_id", school_id)
        .eq("attempt_id", attempt_id)
        .eq("question_id", question_id)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    save_payload = {
        "school_id": school_id,
        "attempt_id": attempt_id,
        "test_id": _normalize(attempt.get("test_id")),
        "question_id": question_id,
        "student_id": _normalize(student.get("id")),
        "response_payload": response_payload,
        "is_marked_for_review": bool(payload.get("is_marked_for_review", False)),
        "answered_at": _utc_now_iso(),
        "metadata": {},
        "is_active": True,
    }
    if existing_rows:
        _table("test_responses").update(save_payload).eq("school_id", school_id).eq(
            "id", _normalize(existing_rows[0].get("id"))
        ).execute()
    else:
        _table("test_responses").insert(save_payload).execute()
    active_responses = _attempt_responses_rows(school_id, attempt_id)
    _table("test_attempts").update({"answered_questions_snapshot": len(active_responses)}).eq(
        "school_id", school_id
    ).eq("id", attempt_id).execute()
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.attempt_saved",
        entity_id=attempt_id,
        payload={"question_id": question_id, "response_saved": True},
    )
    touch_test_session(
        school_id,
        _normalize(attempt.get("test_id")),
        _normalize(student.get("id")),
        _normalize(payload.get("session_key")) or None,
    )
    return get_attempt(school_id, attempt_id)


def submit_attempt(school_id: str, attempt_id: str, profile_id: str, session_key: str | None = None) -> dict[str, Any]:
    from app.services.supabase_account_security import end_test_session

    attempt = _get_attempt_row(school_id, attempt_id)
    student = _get_student_by_profile_id(school_id, profile_id)
    test = _get_test_row(school_id, _normalize(attempt.get("test_id")))
    if _normalize(attempt.get("student_id")) != _normalize(student.get("id")):
        raise HTTPException(status_code=403, detail="Students can submit only their own attempts")
    if _normalize(attempt.get("status")) != "in_progress":
        raise HTTPException(status_code=400, detail="Attempt has already been submitted")
    question_rows = _question_rows_for_test(school_id, _normalize(attempt.get("test_id")))
    question_map = {_normalize(row.get("id")): row for row in question_rows}
    response_rows = _attempt_responses_rows(school_id, attempt_id)

    answered_questions = 0
    correct_answers = 0
    incorrect_answers = 0
    score_obtained = 0.0
    max_score = 0.0
    for question in question_rows:
        max_score += float(question.get("marks") or 0)
    for response_row in response_rows:
        question = question_map.get(_normalize(response_row.get("question_id")))
        if not question:
            continue
        answered_questions += 1
        is_correct, marks_awarded = _score_response(question, _normalize_json_object(response_row.get("response_payload")))
        if is_correct is True:
            correct_answers += 1
        elif is_correct is False:
            incorrect_answers += 1
        score_obtained += marks_awarded
        _table("test_responses").update(
            {
                "is_correct": is_correct,
                "marks_awarded": marks_awarded,
                "evaluated_at": _utc_now_iso(),
            }
        ).eq("school_id", school_id).eq("id", _normalize(response_row.get("id"))).execute()
    total_questions = len(question_rows)
    unanswered_questions = max(total_questions - answered_questions, 0)
    percentage = round((score_obtained / max_score) * 100, 2) if max_score > 0 else 0.0
    submitted_at = _utc_now_iso()
    pass_marks = float(test.get("pass_marks")) if test.get("pass_marks") is not None else None
    passed = score_obtained >= pass_marks if pass_marks is not None else None
    auto_submitted_at = None
    started_at = attempt.get("started_at")
    if started_at and test.get("duration_minutes"):
        started_at_dt = datetime.fromisoformat(str(started_at).replace("Z", "+00:00"))
        if datetime.now(timezone.utc) >= started_at_dt + timedelta(minutes=int(test.get("duration_minutes") or 0)):
            auto_submitted_at = submitted_at
    _table("test_attempts").update(
        {
            "status": "submitted",
            "submitted_at": submitted_at,
            "auto_submitted_at": auto_submitted_at,
            "evaluated_at": submitted_at,
            "answered_questions_snapshot": answered_questions,
        }
    ).eq("school_id", school_id).eq("id", attempt_id).execute()
    existing_result_rows = list(
        _table("test_results")
        .select("*")
        .eq("school_id", school_id)
        .eq("attempt_id", attempt_id)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
        .data
        or []
    )
    result_payload = {
        "school_id": school_id,
        "attempt_id": attempt_id,
        "test_id": _normalize(attempt.get("test_id")),
        "student_id": _normalize(student.get("id")),
        "status": "evaluated",
        "total_questions": total_questions,
        "attempted_questions": answered_questions,
        "correct_answers": correct_answers,
        "incorrect_answers": incorrect_answers,
        "unanswered_questions": unanswered_questions,
        "score_obtained": round(score_obtained, 2),
        "max_score": round(max_score, 2),
        "percentage": percentage,
        "published_at": submitted_at,
        "metadata": {
            "auto_evaluated": True,
            "passed": passed,
            "pass_marks": pass_marks,
            "auto_submitted": bool(auto_submitted_at),
        },
        "is_active": True,
    }
    if existing_result_rows:
        result_id = _normalize(existing_result_rows[0].get("id"))
        _table("test_results").update(result_payload).eq("school_id", school_id).eq("id", result_id).execute()
        _recalculate_result_ranks(school_id, _normalize(attempt.get("test_id")))
        result = get_result(school_id, result_id)
    else:
        response = _table("test_results").insert(result_payload).execute()
        rows = list(response.data or [])
        if not rows:
            raise HTTPException(status_code=500, detail="Result creation failed during attempt submission")
        created_result_id = _normalize(rows[0].get("id"))
        _recalculate_result_ranks(school_id, _normalize(attempt.get("test_id")))
        result = get_result(school_id, created_result_id)
    _log_audit_entry(
        school_id=school_id,
        profile_id=profile_id,
        action="online_tests.attempt_submitted",
        entity_id=attempt_id,
        payload={"test_id": _normalize(attempt.get("test_id")), "result_id": result.get("id")},
    )
    end_test_session(
        school_id,
        _normalize(attempt.get("test_id")),
        _normalize(student.get("id")),
        _normalize(session_key) or None,
        reason="completed",
    )
    return result


def get_result(school_id: str, result_id: str) -> dict[str, Any]:
    return _serialize_result(_get_result_row(school_id, result_id))


def list_results(
    school_id: str,
    *,
    test_id: str | None = None,
    student_id: str | None = None,
    skip: int = 0,
    limit: int = 100,
) -> list[dict[str, Any]]:
    query = _table("test_results").select("*").eq("school_id", school_id).is_("deleted_at", "null")
    if test_id:
        query = query.eq("test_id", test_id)
    if student_id:
        query = query.eq("student_id", student_id)
    rows = list(
        query
        .order("created_at", desc=True)
        .range(max(skip, 0), max(skip, 0) + max(limit, 1) - 1)
        .execute()
        .data
        or []
    )
    return [_serialize_result(dict(row)) for row in rows]


def get_results_analytics(school_id: str | None, *, test_id: str | None = None, global_scope: bool = False) -> dict[str, Any]:
    tests_query = _table("tests").select("*").is_("deleted_at", "null")
    attempts_query = _table("test_attempts").select("*").is_("deleted_at", "null")
    results_query = _table("test_results").select("*").is_("deleted_at", "null")
    if school_id:
        tests_query = tests_query.eq("school_id", school_id)
        attempts_query = attempts_query.eq("school_id", school_id)
        results_query = results_query.eq("school_id", school_id)
    if test_id:
        tests_query = tests_query.eq("id", test_id)
        attempts_query = attempts_query.eq("test_id", test_id)
        results_query = results_query.eq("test_id", test_id)
    test_rows = list(tests_query.execute().data or [])
    attempt_rows = list(attempts_query.execute().data or [])
    result_rows = [dict(row) for row in list(results_query.execute().data or [])]
    percentages = [float(row.get("percentage") or 0) for row in result_rows]
    scores = [float(row.get("score_obtained") or 0) for row in result_rows]
    question_wise_analysis: list[dict[str, Any]] = []
    difficulty_wise_analysis: list[dict[str, Any]] = []
    student_ranking: list[dict[str, Any]] = []

    if school_id and test_id:
        question_rows = _question_rows_for_test(school_id, test_id)
        question_map = {_normalize(item.get("id")): item for item in question_rows}
        response_rows = [
            dict(row)
            for row in list(
                _table("test_responses")
                .select("question_id,is_correct,marks_awarded")
                .eq("school_id", school_id)
                .eq("test_id", test_id)
                .is_("deleted_at", "null")
                .execute()
                .data
                or []
            )
        ]
        question_stats: dict[str, dict[str, Any]] = {}
        difficulty_stats: dict[str, dict[str, Any]] = {}
        for response in response_rows:
            question_id = _normalize(response.get("question_id"))
            question = question_map.get(question_id)
            if not question:
                continue
            stats = question_stats.setdefault(
                question_id,
                {
                    "question_id": question_id,
                    "prompt_text": _normalize(question.get("prompt_text")),
                    "difficulty_level": _normalize(question.get("difficulty_level")) or "medium",
                    "attempts": 0,
                    "correct": 0,
                    "incorrect": 0,
                    "average_marks": 0.0,
                },
            )
            stats["attempts"] += 1
            if response.get("is_correct") is True:
                stats["correct"] += 1
            elif response.get("is_correct") is False:
                stats["incorrect"] += 1
            stats["average_marks"] += float(response.get("marks_awarded") or 0)
        for stats in question_stats.values():
            attempts = max(int(stats["attempts"]), 1)
            stats["average_marks"] = round(float(stats["average_marks"]) / attempts, 2)
            stats["correct_rate"] = round((int(stats["correct"]) / attempts) * 100, 2)
            question_wise_analysis.append(stats)
            difficulty_bucket = difficulty_stats.setdefault(
                str(stats["difficulty_level"]),
                {"difficulty_level": str(stats["difficulty_level"]), "questions": 0, "attempts": 0, "correct": 0},
            )
            difficulty_bucket["questions"] += 1
            difficulty_bucket["attempts"] += attempts
            difficulty_bucket["correct"] += int(stats["correct"])
        for stats in difficulty_stats.values():
            attempts = max(int(stats["attempts"]), 1)
            stats["correct_rate"] = round((int(stats["correct"]) / attempts) * 100, 2)
            difficulty_wise_analysis.append(stats)

        ranked_results = sorted(result_rows, key=lambda row: (float(row.get("percentage") or 0), float(row.get("score_obtained") or 0)), reverse=True)
        student_ids = [_normalize(row.get("student_id")) for row in ranked_results]
        student_map = _student_rows_by_ids(school_id, student_ids)
        for index, row in enumerate(ranked_results[:20], start=1):
            student = student_map.get(_normalize(row.get("student_id")), {})
            student_ranking.append(
                {
                    "rank": index,
                    "student_id": _normalize(row.get("student_id")),
                    "student_name": _normalize(student.get("full_name")) or f"Student {_normalize(row.get('student_id'))[:8]}",
                    "batch_id": _normalize(student.get("batch_id")) or None,
                    "percentage": round(float(row.get("percentage") or 0), 2),
                    "score_obtained": round(float(row.get("score_obtained") or 0), 2),
                    "max_score": round(float(row.get("max_score") or 0), 2),
                }
            )
    return {
        "scope": "global" if global_scope else "school",
        "school_id": school_id,
        "test_id": test_id,
        "total_tests": len(test_rows),
        "total_attempts": len(attempt_rows),
        "completed_attempts": len(
            [row for row in attempt_rows if _normalize(row.get("status")) in {"submitted", "evaluated"}]
        ),
        "evaluated_results": len(result_rows),
        "average_score": round(sum(scores) / len(scores), 2) if scores else 0.0,
        "average_percentage": round(sum(percentages) / len(percentages), 2) if percentages else 0.0,
        "highest_score": round(max(scores), 2) if scores else 0.0,
        "lowest_score": round(min(scores), 2) if scores else 0.0,
        "published_results": len([row for row in result_rows if row.get("published_at")]),
        "question_wise_analysis": question_wise_analysis,
        "difficulty_wise_analysis": difficulty_wise_analysis,
        "student_ranking": student_ranking,
    }

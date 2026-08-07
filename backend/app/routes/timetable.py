"""Timetable management routes (Supabase-native)."""

from __future__ import annotations

import asyncio
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.middleware.auth import get_authenticated_actor_context, require_permissions
from app.middleware.tenant_context import TenantContext, get_tenant_context
from app.models import User, UserRole
from app.schemas import (
    ConflictCheckResponse,
    DayOfWeek,
    TimetableEntryCreate,
    TimetableEntryResponse,
    TimetableEntryUpdate,
    TimetableView,
)
from app.services.supabase_admin import get_supabase_admin_client
from app.services.scope_engine import (
    PermissionScopeContext,
    build_scope_context,
    ensure_school_wide_scope,
)
from app.services.supabase_timetable import (
    check_batch_conflicts as check_batch_conflicts_supabase,
    check_room_conflicts as check_room_conflicts_supabase,
    check_teacher_conflicts as check_teacher_conflicts_supabase,
    create_timetable_entry as create_timetable_entry_supabase,
    delete_all_timetable_entries as delete_all_timetable_entries_supabase,
    delete_timetable_entry as delete_timetable_entry_supabase,
    get_timetable_entry as get_timetable_entry_supabase,
    list_timetable_entries as list_timetable_entries_supabase,
    update_timetable_entry as update_timetable_entry_supabase,
)
from app.utils.dashboard_tracing import begin_dashboard_request, finish_dashboard_request

router = APIRouter()
utility_router = APIRouter()

DAYS_ORDER = [
    DayOfWeek.MONDAY,
    DayOfWeek.TUESDAY,
    DayOfWeek.WEDNESDAY,
    DayOfWeek.THURSDAY,
    DayOfWeek.FRIDAY,
    DayOfWeek.SATURDAY,
    DayOfWeek.SUNDAY,
]
DAYS_LABELS = {day.value: day.value.capitalize() for day in DAYS_ORDER}
EXPORT_GROUPINGS = {"day", "teacher", "room", "batch"}
SESSION_MODE_FILTERS = {"all", "offline", "online", "merged"}
BREAK_TEACHER_NAME = "__BREAK_SESSION__"
SELF_STUDY_TEACHER_NAME = "__SELF_STUDY_SESSION__"
DAY_NAMES = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]


def get_day_of_week_from_date(date_str: str) -> str:
    parts = date_str.split("-")
    d = date(int(parts[0]), int(parts[1]), int(parts[2]))
    return DAY_NAMES[d.weekday()]


def require_timetable_manage_access(
    actor: dict[str, str] = Depends(get_authenticated_actor_context),
) -> dict[str, str]:
    if actor.get("role") == UserRole.TEACHER.value:
        raise HTTPException(status_code=403, detail="Teachers can only view their own timetable")
    return actor


def require_timetable_view_scope(
    tenant: TenantContext = Depends(get_tenant_context),
    actor: dict[str, str] = Depends(get_authenticated_actor_context),
    user: User = Depends(require_permissions("timetable.view", "timetable")),
) -> PermissionScopeContext:
    school_id = tenant.school_id
    return build_scope_context(
        user=user,
        actor=actor,
        school_id=school_id,
        permission_key="timetable.view",
        include_staff=True,
    )


def require_timetable_manage_scope(
    tenant: TenantContext = Depends(get_tenant_context),
    actor: dict[str, str] = Depends(get_authenticated_actor_context),
    user: User = Depends(require_permissions("timetable.manage", "timetable")),
) -> PermissionScopeContext:
    school_id = tenant.school_id
    return build_scope_context(
        user=user,
        actor=actor,
        school_id=school_id,
        permission_key="timetable.manage",
        include_staff=True,
    )


def _resolve_actor_teacher_scope(school_id: str, actor: dict[str, Any]) -> tuple[str | None, str]:
    actor_email = str(actor.get("email") or "").strip()
    actor_name = str(actor.get("name") or "").strip()
    if not actor_email and not actor_name:
        return None, actor_name.casefold()

    query = (
        get_supabase_admin_client()
        .table("staff_members")
        .select("id, full_name")
        .eq("school_id", school_id)
        .eq("is_active", True)
        .eq("staff_type", "teaching")
    )
    if actor_email:
        query = query.ilike("email", actor_email)
    else:
        query = query.ilike("full_name", actor_name)
    response = query.limit(1).execute()
    rows = list(response.data or [])
    if rows:
        return str(rows[0].get("id") or "").strip() or None, str(rows[0].get("full_name") or actor_name).strip().casefold()
    return None, actor_name.casefold()


def _filter_rows_for_teacher_actor(rows: list[dict[str, Any]], school_id: str, actor: dict[str, Any]) -> list[dict[str, Any]]:
    if actor.get("role") != UserRole.TEACHER.value:
        return rows
    actor_teacher_id, actor_teacher_name = _resolve_actor_teacher_scope(school_id, actor)
    filtered: list[dict[str, Any]] = []
    for row in rows:
        if actor_teacher_id and str(row.get("teacher_id") or row.get("staff_member_id") or "").strip() == actor_teacher_id:
            filtered.append(row)
            continue
        teacher_name = str(row.get("teacher_name") or "").strip().casefold()
        if actor_teacher_name and teacher_name and teacher_name == actor_teacher_name:
            filtered.append(row)
    return filtered


def _enforce_teacher_entry_scope(entry: dict[str, Any], school_id: str, actor: dict[str, Any]) -> None:
    if actor.get("role") != UserRole.TEACHER.value:
        return
    actor_teacher_id, actor_teacher_name = _resolve_actor_teacher_scope(school_id, actor)
    entry_teacher_id = str(entry.get("teacher_id") or entry.get("staff_member_id") or "").strip()
    entry_teacher_name = str(entry.get("teacher_name") or "").strip().casefold()
    if actor_teacher_id and entry_teacher_id == actor_teacher_id:
        return
    if actor_teacher_name and entry_teacher_name and entry_teacher_name == actor_teacher_name:
        return
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Teachers can only view their own timetable")


def _filter_rows_for_scope(rows: list[dict[str, Any]], context: PermissionScopeContext) -> list[dict[str, Any]]:
    if context.is_school_wide:
        return rows
    staff_member_id = str(context.staff_member_id or "").strip()
    if not staff_member_id:
        return []
    return [
        row
        for row in rows
        if str(row.get("teacher_id") or row.get("staff_member_id") or "").strip() == staff_member_id
    ]


def _enforce_scope_teacher_access(context: PermissionScopeContext, teacher_id: str | None, detail: str) -> None:
    if context.is_school_wide:
        return
    if not context.staff_member_id or str(teacher_id or "").strip() != str(context.staff_member_id).strip():
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=detail)


def _enforce_scope_entry_access(context: PermissionScopeContext, entry: dict[str, Any], detail: str) -> None:
    if context.is_school_wide:
        return
    _enforce_scope_teacher_access(
        context,
        str(entry.get("teacher_id") or entry.get("staff_member_id") or "").strip() or None,
        detail,
    )


def format_time_label(time_value: str) -> str:
    parsed = datetime.strptime(time_value, "%H:%M")
    return parsed.strftime("%I:%M %p")


def format_time_range(start_time: str, end_time: str) -> str:
    return f"{format_time_label(start_time)} TO {format_time_label(end_time)}"


def format_session_type(value: str | None) -> str:
    mapping = {
        "regular_class": "Regular Class",
        "break_time": "Break Time",
        "doubt_session": "Doubt Session",
        "extra_class": "Extra Class",
        "self_study": "Self Study",
    }
    return mapping.get(value or "regular_class", "Regular Class")


def is_break_entry(entry: TimetableView) -> bool:
    subject_value = (entry.subject or "").strip().lower()
    return entry.session_type == "break_time" or subject_value == "break time"


def is_self_study_entry(entry: TimetableView) -> bool:
    subject_value = (entry.subject or "").strip().lower()
    return entry.session_type == "self_study" or subject_value == "self study"


def is_no_teacher_session(session_type: str | None, subject: str | None = None) -> bool:
    normalized_type = (session_type or "").strip().lower()
    normalized_subject = (subject or "").strip().lower()
    return normalized_type in {"break_time", "self_study"} or normalized_subject in {"break time", "self study"}


def build_subject_label(entry: TimetableView) -> str:
    if is_break_entry(entry):
        return "BREAK TIME"
    return (entry.subject or "").strip()


def build_location_label(entry: TimetableView) -> str:
    if is_break_entry(entry):
        return "BREAK TIME"
    return entry.room_name or "-"


def build_teacher_label(entry: TimetableView) -> str:
    if is_break_entry(entry):
        return "BREAK TIME"
    if is_self_study_entry(entry):
        return "SELF STUDY"
    return entry.teacher_name or ""


def build_mode_type_label(entry: TimetableView) -> str:
    if is_break_entry(entry):
        return "BREAK TIME"
    return "Online" if entry.session_mode == "online" else "Offline"


def normalize_session_mode_filter(session_mode_filter: str | None) -> str:
    value = (session_mode_filter or "all").strip().lower()
    return value if value in SESSION_MODE_FILTERS else "all"


def build_export_title(view_by: str, session_mode_filter: str) -> str:
    if session_mode_filter == "merged":
        return f"TIMETABLE EXPORT - MERGED ONLINE/OFFLINE - {view_by.upper()} WISE"
    if session_mode_filter in {"online", "offline"}:
        return f"TIMETABLE EXPORT - {session_mode_filter.upper()} - {view_by.upper()} WISE"
    return f"TIMETABLE EXPORT - {view_by.upper()} WISE"


def split_batch_names(class_name: str) -> list[str]:
    values = [(item or "").strip() for item in (class_name or "").split(",")]
    return [item for item in values if item]


def build_timetable_view(entry: dict[str, Any]) -> TimetableView:
    view = TimetableView(
        id=entry.get("id"),
        day_of_week=entry.get("day_of_week"),
        start_time=entry.get("start_time"),
        end_time=entry.get("end_time"),
        class_name=entry.get("class_name") or "",
        subject=entry.get("subject") or "",
        teacher_name=entry.get("teacher_name") or "",
        teacher_id=entry.get("teacher_id"),
        room_id=entry.get("room_id"),
        room_name=entry.get("room_name") or "",
        session_mode=entry.get("session_mode") or "offline",
        session_type=entry.get("session_type") or "regular_class",
        extra_class_scope=entry.get("extra_class_scope"),
        online_platform=entry.get("online_platform"),
        online_link=entry.get("online_link"),
        online_provider=entry.get("online_provider"),
        meeting_link=entry.get("meeting_link"),
        meeting_id=entry.get("meeting_id"),
        meeting_password=entry.get("meeting_password"),
        recording_url=entry.get("recording_url"),
        notes=entry.get("notes"),
    )
    if is_break_entry(view):
        view.teacher_name = "BREAK TIME"
    elif is_self_study_entry(view):
        view.teacher_name = "SELF STUDY"
    return view


def build_timetable_response(entry: dict[str, Any]) -> TimetableEntryResponse:
    response = TimetableEntryResponse(
        id=entry.get("id"),
        teacher_id=entry.get("teacher_id"),
        room_id=entry.get("room_id"),
        school_id=entry.get("school_id"),
        day_of_week=entry.get("day_of_week"),
        start_time=entry.get("start_time"),
        end_time=entry.get("end_time"),
        class_name=entry.get("class_name") or "",
        subject=entry.get("subject") or "",
        is_active=entry.get("is_active", True),
        created_at=entry.get("created_at"),
        updated_at=entry.get("updated_at"),
        teacher_name=entry.get("teacher_name") or "",
        room_name=entry.get("room_name") or "",
        session_mode=entry.get("session_mode") or "offline",
        session_type=entry.get("session_type") or "regular_class",
        extra_class_scope=entry.get("extra_class_scope"),
        online_platform=entry.get("online_platform"),
        online_link=entry.get("online_link"),
        online_provider=entry.get("online_provider"),
        meeting_link=entry.get("meeting_link"),
        meeting_id=entry.get("meeting_id"),
        meeting_password=entry.get("meeting_password"),
        recording_url=entry.get("recording_url"),
        notes=entry.get("notes"),
    )
    session_type = response.session_type or ""
    subject = (response.subject or "").strip().lower()
    if session_type == "break_time" or subject == "break time":
        response.teacher_name = "Break Time"
    elif session_type == "self_study" or subject == "self study":
        response.teacher_name = "Self Study"
    return response


def coerce_timetable_views(entries: list[dict[str, Any] | TimetableView]) -> list[TimetableView]:
    normalized: list[TimetableView] = []
    for entry in entries:
        if isinstance(entry, TimetableView):
            normalized.append(entry)
        else:
            normalized.append(TimetableView(**entry))
    return normalized


def group_entries_for_export(entries: list[TimetableView], view_by: str, session_mode_filter: str = "all") -> list[dict[str, Any]]:
    grouped: dict[str, list[TimetableView]] = defaultdict(list)
    titles: dict[str, str] = {}
    normalized_mode_filter = normalize_session_mode_filter(session_mode_filter)

    for entry in entries:
        entry_mode = (entry.session_mode or "offline").upper()
        prefix = f"{entry_mode} | " if normalized_mode_filter == "merged" else ""
        mode_key_prefix = f"{entry.session_mode or 'offline'}::" if normalized_mode_filter == "merged" else ""

        if view_by == "teacher":
            key = f"{mode_key_prefix}{entry.teacher_name}"
            title = f"{prefix}{entry.teacher_name.upper()}"
        elif view_by == "room":
            room_label = entry.room_name or "NO ROOM"
            key = f"{mode_key_prefix}{room_label}"
            title = f"{prefix}{room_label.upper()}"
        elif view_by == "batch":
            key = f"{mode_key_prefix}{entry.class_name}"
            title = f"{prefix}{entry.class_name.upper()}"
        else:
            key = f"{entry.session_mode or 'offline'}-{entry.day_of_week.value}" if normalized_mode_filter == "merged" else entry.day_of_week.value
            title = f"{prefix}{DAYS_LABELS[entry.day_of_week.value].upper()} TIMETABLE"

        grouped[key].append(entry)
        titles[key] = title

    sections: list[dict[str, Any]] = []
    for key in sorted(grouped.keys()):
        section_entries = sorted(
            grouped[key],
            key=lambda item: (item.day_of_week.value, item.start_time, item.teacher_name.lower(), item.class_name.lower()),
        )
        sections.append({
            "key": key,
            "title": titles[key],
            "entries": section_entries,
        })
    return sections


def create_timetable_excel(entries: list[TimetableView], view_by: str, session_mode_filter: str = "all") -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Timetable Export"
    worksheet.freeze_panes = "A2"

    thin = Side(style="thin", color="1E3A8A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="00A3E0")
    time_fill = PatternFill("solid", fgColor="F4C542")
    row_fill = PatternFill("solid", fgColor="E8FF72")

    worksheet.merge_cells("A1:G1")
    worksheet["A1"] = build_export_title(view_by, normalize_session_mode_filter(session_mode_filter))
    worksheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    worksheet["A1"].fill = title_fill
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    row_cursor = 3
    sections = group_entries_for_export(entries, view_by, session_mode_filter)
    for section in sections:
        worksheet.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=7)
        cell = worksheet.cell(row=row_cursor, column=1)
        cell.value = section["title"]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        row_cursor += 1

        headers = ["TIMINGS", "DAY", "BATCH / CLASS", "TEACHER", "SUBJECT", "MODE TYPE", "ROOM"]
        for col_index, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=row_cursor, column=col_index)
            header_cell.value = header
            header_cell.font = Font(bold=True)
            header_cell.fill = time_fill if col_index == 1 else title_fill
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.border = border
        row_cursor += 1

        for entry in section["entries"]:
            values = [
                format_time_range(entry.start_time, entry.end_time),
                DAYS_LABELS[entry.day_of_week.value],
                entry.class_name,
                build_teacher_label(entry),
                build_subject_label(entry),
                build_mode_type_label(entry),
                build_location_label(entry),
            ]
            for col_index, value in enumerate(values, start=1):
                data_cell = worksheet.cell(row=row_cursor, column=col_index)
                data_cell.value = value
                data_cell.fill = time_fill if col_index == 1 else row_fill
                data_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                data_cell.border = border
            row_cursor += 1

        row_cursor += 1

    widths = {"A": 22, "B": 14, "C": 28, "D": 26, "E": 24, "F": 16, "G": 18}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def create_timetable_pdf(entries: list[TimetableView], view_by: str, session_mode_filter: str = "all") -> BytesIO:
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Table, TableStyle, Spacer, KeepTogether
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.styles import ParagraphStyle
    from app.utils.pdf_base import (
        ReportPdfBuilder,
        build_shared_styles,
        NAVY, SLATE_700, SLATE_500, SLATE_300, SLATE_200, SLATE_100, SLATE_50,
        WHITE, DARK_TEXT, MEDIUM_TEXT,
        make_paragraph, safe_pdf_text, fmt_timestamp,
    )

    buffer = BytesIO()
    cm = 0.35 * inch
    builder = ReportPdfBuilder(
        buffer,
        pagesize=landscape(A4),
        left_margin=cm, right_margin=cm,
        top_margin=1.2 * inch,
        bottom_margin=0.7 * inch,
        title="Timetable Export",
        author="Sitting Plan System",
    )

    pw, ph = landscape(A4)

    def _timetable_header(canv, ctx):
        canv.setFillColor(NAVY)
        canv.setFont("Helvetica-Bold", 14)
        canv.drawString(cm, ph - 22, "TIMETABLE")
        canv.setFillColor(SLATE_700)
        canv.setFont("Helvetica", 9)
        canv.drawString(cm, ph - 38, safe_pdf_text(ctx.get("title", "")))
        canv.setFillColor(SLATE_500)
        canv.setFont("Helvetica", 7.5)
        canv.drawString(cm, ph - 50, f"Generated: {fmt_timestamp()}")
        canv.setStrokeColor(SLATE_300)
        canv.setLineWidth(0.5)
        canv.line(cm, ph - 56, pw - cm, ph - 56)

    styles = build_shared_styles()
    title_text = build_export_title(view_by, normalize_session_mode_filter(session_mode_filter))

    # Custom section style (flat text color, no bg — B&W friendly)
    section_heading = ParagraphStyle(
        "TTSection",
        fontName="Helvetica-Bold",
        fontSize=10.5,
        leading=13,
        textColor=NAVY,
        spaceBefore=8,
        spaceAfter=4,
    )

    # Time column style - light slate bg (B&W friendly)
    time_style = ParagraphStyle(
        "TTTime",
        parent=styles["table_body_center"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
    )

    builder.add_title(title_text)
    builder.add_spacer(0.08 * inch)

    sections = group_entries_for_export(entries, view_by, session_mode_filter)
    if not sections:
        builder.add_small_note("No timetable entries found for the selected criteria.")
        return builder.build(header_context={"title": title_text})

    for section in sections:
        heading = Paragraph(safe_pdf_text(str(section["title"])), section_heading)
        table_rows = [["TIMINGS", "DAY", "BATCH / CLASS", "TEACHER", "SUBJECT", "MODE TYPE", "ROOM"]]
        for entry in section["entries"]:
            table_rows.append([
                format_time_range(entry.start_time, entry.end_time),
                DAYS_LABELS[entry.day_of_week.value],
                entry.class_name,
                build_teacher_label(entry),
                build_subject_label(entry),
                build_mode_type_label(entry),
                build_location_label(entry),
            ])

        # Professional colour scheme: navy header, light slate time col, white/gray rows
        table = Table(table_rows, colWidths=[100, 60, 115, 105, 105, 70, 80], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("BACKGROUND", (0, 1), (0, -1), SLATE_50),
            ("BACKGROUND", (1, 1), (-1, -1), WHITE),
            ("ROWBACKGROUNDS", (1, 2), (-1, -1), [WHITE, SLATE_50]),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("GRID", (0, 0), (-1, -1), 0.4, SLATE_200),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
            ("TOPPADDING", (0, 0), (-1, 0), 7),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        builder.add_keep_together([heading, table])
        builder.add_spacer(0.08 * inch)

    return builder.build(header_context={"title": title_text})


@router.post("", response_model=TimetableEntryResponse)
async def create_timetable_entry(
    entry: TimetableEntryCreate,
    tenant: TenantContext = Depends(get_tenant_context),
    scope_context: PermissionScopeContext = Depends(require_timetable_manage_scope),
):
    school_id = tenant.school_id
    _enforce_scope_teacher_access(scope_context, str(entry.teacher_id) if entry.teacher_id else None, "You can only create timetable entries for your assigned timetable scope")
    result = create_timetable_entry_supabase(school_id, entry.model_dump())
    return build_timetable_response(result)


@router.get("", response_model=list[TimetableView])
async def list_timetable_entries(
    tenant: TenantContext = Depends(get_tenant_context),
    day_of_week: DayOfWeek | None = None,
    teacher_id: str | int | None = None,
    class_name: str | None = None,
    room_id: str | int | None = None,
    reference_date: date | None = Query(default=None),
    scope_context: PermissionScopeContext = Depends(require_timetable_view_scope),
):
    school_id = tenant.school_id
    if teacher_id and not scope_context.is_school_wide:
        _enforce_scope_teacher_access(scope_context, str(teacher_id), "You can only view your assigned timetable")
    rows = await asyncio.get_event_loop().run_in_executor(
        None,
        lambda: list_timetable_entries_supabase(
            school_id,
            day_of_week=day_of_week.value if day_of_week else None,
            teacher_id=str(teacher_id) if teacher_id else None,
            class_name=class_name,
            room_id=str(room_id) if room_id else None,
            reference_date=reference_date.isoformat() if reference_date else None,
        ),
    )
    rows = _filter_rows_for_scope(rows, scope_context)
    return [build_timetable_view(r) for r in rows]


@router.get("/count")
async def get_timetable_entries_count(
    response: Response,
    tenant: TenantContext = Depends(get_tenant_context),
    scope_context: PermissionScopeContext = Depends(require_timetable_view_scope),
):
    school_id = tenant.school_id
    trace = begin_dashboard_request("timetable_count", school_id)
    response.headers["X-Dashboard-Request-Id"] = str(trace["request_id"])
    try:
        entries = await asyncio.to_thread(list_timetable_entries_supabase, school_id)
        entries = _filter_rows_for_scope(entries, scope_context)
        finish_dashboard_request(trace, cache_status="service_cache_possible", execution_path="list_and_filter")
        return len(entries)
    except Exception as exc:
        finish_dashboard_request(trace, cache_status="service_cache_possible", execution_path="error", error=str(exc)[:200])
        raise


@router.get("/export")
async def export_timetable(
    export_format: str = Query(..., pattern="^(excel|pdf)$"),
    view_by: str = Query(default="day", pattern="^(day|teacher|room|batch)$"),
    session_mode_filter: str = Query(default="all", pattern="^(all|offline|online|merged)$"),
    tenant: TenantContext = Depends(get_tenant_context),
    day_of_week: DayOfWeek | None = Query(default=None),
    teacher_id: str | int | None = Query(default=None),
    room_id: str | int | None = Query(default=None),
    batch_name: str | None = Query(default=None),
    scope_context: PermissionScopeContext = Depends(require_timetable_view_scope),
):
    school_id = tenant.school_id
    if teacher_id and not scope_context.is_school_wide:
        _enforce_scope_teacher_access(scope_context, str(teacher_id), "You can only export your assigned timetable")
    entries = coerce_timetable_views(_filter_rows_for_scope(list_timetable_entries_supabase(
        school_id,
        day_of_week=day_of_week.value if day_of_week else None,
        teacher_id=str(teacher_id) if teacher_id else None,
        class_name=batch_name if view_by == "batch" else None,
        room_id=str(room_id) if room_id else None,
    ), scope_context))
    if not entries:
        raise HTTPException(status_code=404, detail="No timetable entries found for export")

    if export_format == "excel":
        buffer = create_timetable_excel(entries, view_by, session_mode_filter)
        filename = f"timetable-{session_mode_filter}-{view_by}-wise.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        buffer = create_timetable_pdf(entries, view_by, session_mode_filter)
        filename = f"timetable-{session_mode_filter}-{view_by}-wise.pdf"
        media_type = "application/pdf"

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{entry_id}", response_model=TimetableEntryResponse)
async def get_timetable_entry(
    entry_id: str,
    tenant: TenantContext = Depends(get_tenant_context),
    scope_context: PermissionScopeContext = Depends(require_timetable_view_scope),
):
    school_id = tenant.school_id
    result = get_timetable_entry_supabase(school_id, entry_id)
    _enforce_scope_entry_access(scope_context, result, "You can only view timetable entries in your assigned scope")
    return build_timetable_response(result)


@router.put("/{entry_id}", response_model=TimetableEntryResponse)
async def update_timetable_entry(
    entry_id: str,
    entry_update: TimetableEntryUpdate,
    tenant: TenantContext = Depends(get_tenant_context),
    scope_context: PermissionScopeContext = Depends(require_timetable_manage_scope),
):
    school_id = tenant.school_id
    current = get_timetable_entry_supabase(school_id, entry_id)
    _enforce_scope_entry_access(scope_context, current, "You can only update timetable entries in your assigned scope")
    next_teacher_id = str(entry_update.teacher_id) if entry_update.teacher_id else str(current.get("teacher_id") or "")
    _enforce_scope_teacher_access(scope_context, next_teacher_id, "You can only assign timetable entries to yourself")
    result = update_timetable_entry_supabase(school_id, entry_id, entry_update.model_dump(exclude_unset=True))
    return build_timetable_response(result)


@router.delete("/{entry_id}")
async def delete_timetable_entry(
    entry_id: str,
    tenant: TenantContext = Depends(get_tenant_context),
    scope_context: PermissionScopeContext = Depends(require_timetable_manage_scope),
):
    school_id = tenant.school_id
    current = get_timetable_entry_supabase(school_id, entry_id)
    _enforce_scope_entry_access(scope_context, current, "You can only delete timetable entries in your assigned scope")
    return delete_timetable_entry_supabase(school_id, entry_id)


@router.delete("")
async def delete_all_timetable_entries(
    tenant: TenantContext = Depends(get_tenant_context),
    is_admin: bool = Query(default=False),
    scope_context: PermissionScopeContext = Depends(require_timetable_manage_scope),
):
    school_id = tenant.school_id
    ensure_school_wide_scope(scope_context, "Only school-wide timetable access can bulk delete timetable entries")
    if not is_admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admin can delete all timetable entries")
    return delete_all_timetable_entries_supabase(school_id)


@router.post("/check-conflict", response_model=ConflictCheckResponse)
async def check_conflict(
    teacher_id: str | int = Body(...),
    day_of_week: DayOfWeek = Body(...),
    start_time: str = Body(...),
    end_time: str = Body(...),
    room_id: str | int | None = Body(default=None),
    class_name: str | None = Body(default=None),
    exclude_entry_id: str | int = Body(default=None),
    tenant: TenantContext = Depends(get_tenant_context),
    scope_context: PermissionScopeContext = Depends(require_timetable_manage_scope),
):
    school_id = tenant.school_id
    _enforce_scope_teacher_access(scope_context, str(teacher_id), "You can only validate conflicts for your assigned timetable")
    conflicts = check_teacher_conflicts_supabase(
        school_id,
        str(teacher_id),
        day_of_week.value,
        start_time,
        end_time,
        exclude_entry_id=str(exclude_entry_id) if exclude_entry_id else None,
    )
    if conflicts:
        return ConflictCheckResponse(
            has_conflict=True,
            conflicting_entries=[
                get_timetable_entry_supabase(school_id, str(c["id"]))
                for c in conflicts
            ],
            message="Conflict detected: Teacher is already assigned during this time slot",
        )
    if room_id:
        room_conflicts = check_room_conflicts_supabase(
            school_id,
            str(room_id),
            day_of_week.value,
            start_time,
            end_time,
            exclude_entry_id=str(exclude_entry_id) if exclude_entry_id else None,
        )
        if room_conflicts:
            return ConflictCheckResponse(
                has_conflict=True,
                conflicting_entries=[
                    get_timetable_entry_supabase(school_id, str(c["id"]))
                    for c in room_conflicts
                ],
                message="Conflict detected: Room is already assigned during this time slot",
            )
    if class_name:
        batch_conflicts = check_batch_conflicts_supabase(
            school_id,
            class_name,
            day_of_week.value,
            start_time,
            end_time,
            exclude_entry_id=str(exclude_entry_id) if exclude_entry_id else None,
        )
        if batch_conflicts:
            return ConflictCheckResponse(
                has_conflict=True,
                conflicting_entries=[
                    get_timetable_entry_supabase(school_id, str(c["id"]))
                    for c in batch_conflicts
                ],
                message="Conflict detected: Batch/Class already has a timetable entry during this time slot",
            )
    return ConflictCheckResponse(has_conflict=False, message="No conflicts detected")


@utility_router.get("/template")
async def download_timetable_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable Template"
    headers = ["Date", "Day", "Teacher", "Batch/Class", "Subject", "Start Time", "End Time", "Room", "Mode"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 12
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=timetable_template.xlsx"},
    )


@utility_router.post("/upload")
async def upload_timetable_excel(
    file: UploadFile = File(...),
    tenant: TenantContext = Depends(get_tenant_context),
    scope_context: PermissionScopeContext = Depends(require_timetable_manage_scope),
):
    school_id = tenant.school_id
    ensure_school_wide_scope(scope_context, "Only school-wide timetable access can import timetable workbooks")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel file parse nahi ho paaya: {e}")
    ws = wb.active
    try:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel rows parse nahi ho paaya: {e}")
    created = []
    errors = []
    supabase = get_supabase_admin_client()
    teacher_cache: dict[str, str | None] = {}
    room_cache: dict[str, str | None] = {}
    teacher_subject_cache: dict[str, str] = {}

    try:
        try:
            acad = get_supabase_admin_client()
            assignments_resp = (
                acad.schema("academic").table("staff_subject_assignments")
                .select("staff_member_id, subject_id")
                .eq("school_id", school_id)
                .eq("is_active", True)
                .execute()
            )
            for a in (assignments_resp.data or []):
                sm_id = str(a.get("staff_member_id") or "")
                subj_id = str(a.get("subject_id") or "")
                if sm_id and subj_id:
                    teacher_subject_cache[sm_id] = subj_id
        except Exception as e:
            errors.append(f"Subject assignments load failed: {e}")

        subject_name_cache: dict[str, str] = {}
        try:
            subjects_resp = (
                supabase.table("subjects")
                .select("id, name")
                .eq("school_id", school_id)
                .execute()
            )
            for s in (subjects_resp.data or []):
                subject_name_cache[str(s.get("id"))] = str(s.get("name") or "")
        except Exception as e:
            errors.append(f"Subjects load failed: {e}")

        def resolve_teacher_id(name: str) -> str | None:
            if not name or not name.strip():
                return None
            key = name.strip().lower()
            if key in teacher_cache:
                return teacher_cache[key]
            response = (
                supabase.table("staff_members")
                .select("id")
                .ilike("full_name", f"%{name.strip()}%")
                .eq("school_id", school_id)
                .limit(1)
                .execute()
            )
            data = response.data if isinstance(response.data, list) else []
            result = str(data[0]["id"]) if data else None
            teacher_cache[key] = result
            return result

        def resolve_room_id(name: str) -> str | None:
            if not name or not name.strip():
                return None
            name_key = name.strip().lower()
            if name_key in room_cache:
                return room_cache[name_key]
            response = (
                supabase.table("rooms")
                .select("id")
                .ilike("name", f"%{name.strip()}%")
                .eq("school_id", school_id)
                .limit(1)
                .execute()
            )
            data = response.data if isinstance(response.data, list) else []
            result = str(data[0]["id"]) if data else None
            room_cache[name_key] = result
            return result

        payloads: list[dict[str, Any]] = []
        for idx, row in enumerate(rows, 2):
            date_val, day_val, teacher_name, batch_val, subject_val, start_time, end_time, room_name, mode_val = (
                (str(v).strip() if v else "") for v in (row + (None,) * (9 - len(row)))[:9]
            )
            if not teacher_name or not batch_val or not start_time or not end_time:
                errors.append(f"Row {idx}: Teacher, Batch, Start Time, End Time required")
                continue
            if not date_val:
                errors.append(f"Row {idx}: Date required")
                continue

            teacher_id = resolve_teacher_id(teacher_name)
            if not teacher_id:
                errors.append(f"Row {idx}: Teacher '{teacher_name}' system mein nahi mila")
                continue
            room_id = resolve_room_id(room_name) if room_name else None
            day_of_week = day_val.lower() if day_val else get_day_of_week_from_date(date_val)

            if not subject_val or subject_val.strip() == "":
                detected_subject_id = teacher_subject_cache.get(teacher_id)
                if detected_subject_id:
                    subject_val = subject_name_cache.get(detected_subject_id, "General")
                else:
                    try:
                        teacher_info = (
                            supabase.table("staff_members")
                            .select("department")
                            .eq("id", teacher_id)
                            .single()
                            .execute()
                        )
                        dept = (teacher_info.data or {}).get("department") or ""
                        subject_val = dept if dept.strip() else "General"
                    except Exception:
                        subject_val = "General"

            payloads.append({
                "teacher_id": teacher_id,
                "room_id": room_id,
                "school_id": school_id,
                "day_of_week": day_of_week,
                "start_time": start_time[:5],
                "end_time": end_time[:5],
                "class_name": batch_val,
                "subject": subject_val or "General",
                "session_mode": mode_val.lower() if mode_val in ("offline", "online") else "offline",
                "session_type": "regular_class",
                "start_date": date_val[:10],
            })

        for idx, p in enumerate(payloads):
            try:
                result = create_timetable_entry_supabase(school_id, p)
                created.append(result)
            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

    except Exception as e:
        errors.append(f"Upload processing failed: {e}")

    return {
        "created": len(created),
        "errors": errors,
        "entries": created,
    }

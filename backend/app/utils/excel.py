"""
Excel import/export utilities
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def default_academic_session() -> str:
    now = datetime.now()
    start_year = now.year if now.month >= 4 else now.year - 1
    end_year = start_year + 1
    return f"Apr {start_year} - Mar {end_year}"


def _normalize_header(value: str) -> str:
    import re

    return re.sub(r"[^a-z0-9]", "", value.strip().lower())


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_excel_date(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    raw = str(value).strip()
    if not raw:
        return ""

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return raw


def _parse_bool(value: object) -> bool:
    normalized = _clean_text(value).lower()
    if not normalized:
        return False
    return normalized in {
        "1",
        "true",
        "yes",
        "y",
        "required",
        "hostel",
        "boarding",
        "residential",
        "boarder",
    }


def _first_non_empty(*values: object) -> str:
    for value in values:
        normalized = _clean_text(value)
        if normalized:
            return normalized
    return ""


STUDENT_TEMPLATE_HEADERS = [
    "SR. NO",
    "ADMISSION ID",
    "ACADEMIC SESSION",
    "PROGRAM",
    "COURSE",
    "MANAGED BATCH",
    "CLASS",
    "SECTION",
    "ROLL NO",
    "FIRST NAME",
    "MIDDLE NAME",
    "LAST NAME",
    "LOCAL NAME",
    "DATE OF BIRTH",
    "AGE AS OF TODAY",
    "EMAIL",
    "GENDER",
    "PHONE",
    "ADMISSION DATE",
    "FEE SCHEDULE",
    "TC NUMBER",
    "PREVIOUS SCHOOL",
    "PREVIOUS EXAM",
    "PREVIOUS BOARD",
    "PREVIOUS PERCENTAGE",
    "PREVIOUS TOTAL MARKS",
    "PREVIOUS AVERAGE",
    "FATHER NAME",
    "FATHER MOBILE",
    "FATHER OCCUPATION",
    "MOTHER NAME",
    "MOTHER MOBILE",
    "MOTHER OCCUPATION",
    "CATEGORY",
    "SUB CATEGORY",
    "SIBLING NAME",
    "SIBLING SCHOOL",
    "GUARDIAN NAME",
    "GUARDIAN RELATION",
    "GUARDIAN MOBILE",
    "GUARDIAN ADDRESS",
    "EMERGENCY NAME",
    "EMERGENCY MOBILE",
    "PRIORITY CONTACT",
    "ADDRESS LINE 1",
    "ADDRESS LINE 2",
    "CITY",
    "STATE",
    "COUNTRY",
    "PINCODE",
    "REGION",
    "REFERENCE NAME",
    "REFERENCE NUMBER",
    "REFERENCE REMARK",
    "ADMISSION TYPE",
    "BOARDING TYPE",
    "SPECIAL NEEDS",
    "AVAILING MESS FACILITY",
    "HOSTEL REQUIRED",
    "PREFERRED HOSTEL",
    "HOSTEL REQUEST NOTE",
    "PICKUP ENABLED",
    "DROP ENABLED",
    "TRANSPORT MONTH",
    "TRANSPORT ROUTE",
    "TRANSPORT STOP",
    "ROOM NO",
]


STUDENT_HEADER_ALIASES = {
    "SR. NO": ["srno", "slno", "sno", "serialno", "serial number", "sl.no."],
    "ADMISSION ID": [
        "admissionid",
        "admission id",
        "admission no",
        "admission number",
        "admission",
        "* admission id",
    ],
    "ACADEMIC SESSION": [
        "academicsession",
        "academic session",
        "session",
        "academic year",
        "year session",
        "academicyear",
    ],
    "PROGRAM": ["program", "stream", "school group"],
    "COURSE": ["course", "exam course", "program course", "entrance exam name"],
    "MANAGED BATCH": ["managedbatch", "managed batch", "batch", "admitted class"],
    "CLASS": ["class", "* class", "class name", "grade", "standard"],
    "SECTION": ["section", "* section", "division"],
    "ROLL NO": [
        "rollno",
        "roll number",
        "rollnumber",
        "roll",
        "registration number",
        "registration no",
    ],
    "FIRST NAME": ["firstname", "first name", "* first name", "student first name"],
    "MIDDLE NAME": ["middlename", "middle name"],
    "LAST NAME": ["lastname", "last name", "student last name"],
    "LOCAL NAME": ["localname", "first name local", "last name local"],
    "DATE OF BIRTH": ["dateofbirth", "date of birth", "dob"],
    "AGE AS OF TODAY": ["ageasoftoday", "age as of today", "age as on", "age"],
    "EMAIL": ["email", "email id", "student email", "mail"],
    "GENDER": ["gender", "sex"],
    "PHONE": ["phone", "student phone number", "student mobile", "mobile", "phone number"],
    "ADMISSION DATE": [
        "admissiondate",
        "date of admission",
        "dateofadmissionddmmyyyy",
        "dateofadmission",
    ],
    "FEE SCHEDULE": ["feeschedule", "fee schedule", "fee schedule name"],
    "TC NUMBER": ["tcnumber", "tc number"],
    "PREVIOUS SCHOOL": ["previousschool", "previous school"],
    "PREVIOUS EXAM": ["previousexam", "previous exam", "previous exam passed"],
    "PREVIOUS BOARD": ["previousboard", "previous board", "board"],
    "PREVIOUS PERCENTAGE": ["previouspercentage", "previous percentage", "pcm %"],
    "PREVIOUS TOTAL MARKS": ["previoustotalmarks", "previous total marks", "total marks", "pcm marks"],
    "PREVIOUS AVERAGE": ["previousaverage", "previous average", "average"],
    "FATHER NAME": ["fathername", "father name"],
    "FATHER MOBILE": ["fathermobile", "father mobile", "* father mobile number", "father mobile number"],
    "FATHER OCCUPATION": ["fatheroccupation", "father occupation"],
    "MOTHER NAME": ["mothername", "mother name"],
    "MOTHER MOBILE": ["mothermobile", "mother mobile", "mother mobile number"],
    "MOTHER OCCUPATION": ["motheroccupation", "mother occupation"],
    "CATEGORY": ["category", "category (general / obc / sc / st)", "caste"],
    "SUB CATEGORY": ["subcategory", "sub category", "sub caste"],
    "SIBLING NAME": ["siblingname", "sibling 1 name", "sibling name"],
    "SIBLING SCHOOL": ["siblingschool", "sibling school", "sibling 1 school/working"],
    "GUARDIAN NAME": ["guardianname", "guardian name"],
    "GUARDIAN RELATION": ["guardianrelation", "guardian relationship"],
    "GUARDIAN MOBILE": ["guardianmobile", "guardian mobile number", "guardian mobile"],
    "GUARDIAN ADDRESS": ["guardianaddress", "guardian address"],
    "EMERGENCY NAME": ["emergencyname", "emergency name"],
    "EMERGENCY MOBILE": ["emergencymobile", "emergency number", "emergency mobile"],
    "PRIORITY CONTACT": [
        "prioritycontact",
        "priority to contact for school matters",
        "priority contact",
    ],
    "ADDRESS LINE 1": ["addressline1", "address 1", "address1", "house number", "location"],
    "ADDRESS LINE 2": ["addressline2", "address 2", "address2"],
    "CITY": ["city", "city / country"],
    "STATE": ["state"],
    "COUNTRY": ["country", "nationality"],
    "PINCODE": ["pincode", "pin code", "zip"],
    "REGION": ["region"],
    "REFERENCE NAME": ["referencename", "reference name"],
    "REFERENCE NUMBER": ["referencenumber", "reference number"],
    "REFERENCE REMARK": ["referenceremark", "reference remark"],
    "ADMISSION TYPE": ["admissiontype", "admission type"],
    "BOARDING TYPE": ["boardingtype", "boarding type"],
    "SPECIAL NEEDS": [
        "specialneeds",
        "special needs",
        "allergymedicalconditiondescription",
        "special education needs description",
        "does the student have any special educational needs ?",
        "does the student have any physical health limitation ?",
        "does the student have any illnesses or medical history since birth ?",
        "does the student have any other medical concerns or conditions ?",
        "other information description",
        "other relevant information ?",
        "special needs description",
        "remarks",
    ],
    "AVAILING MESS FACILITY": ["availingmessfacility", "availing mess facility"],
    "HOSTEL REQUIRED": ["hostelrequired", "hostel required"],
    "PREFERRED HOSTEL": ["preferredhostel", "preferred hostel"],
    "HOSTEL REQUEST NOTE": ["hostelrequestnote", "hostel request note", "reason"],
    "PICKUP ENABLED": ["pickupenabled", "pickup enabled"],
    "DROP ENABLED": ["dropenabled", "drop enabled"],
    "TRANSPORT MONTH": ["transportmonth", "transport month"],
    "TRANSPORT ROUTE": ["transportroute", "transport route", "route name"],
    "TRANSPORT STOP": ["transportstop", "transport stop", "stop name"],
    "ROOM NO": ["roomno", "room number", "room no", "room"],
}


def parse_student_excel(file_content: bytes) -> Tuple[List[Dict], List[Dict]]:
    """
    Parse student data from Excel file.

    Supports both the app template headers and the existing school workbook headers.
    """
    valid_students: List[Dict] = []
    errors: List[Dict] = []

    alias_to_header: dict[str, str] = {}
    for canonical, aliases in STUDENT_HEADER_ALIASES.items():
        alias_to_header[_normalize_header(canonical)] = canonical
        for alias in aliases:
            alias_to_header[_normalize_header(alias)] = canonical

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        worksheet = workbook.active

        if worksheet.max_row < 2:
            errors.append({"row": 0, "error": "Excel file appears to be empty."})
            return valid_students, errors

        actual_headers = []
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col_idx).value
            actual_headers.append(str(cell_value).strip() if cell_value is not None else "")

        header_map: dict[str, int] = {}
        for col_idx, actual_header in enumerate(actual_headers, start=1):
            normalized_actual = _normalize_header(actual_header)
            if not normalized_actual:
                continue

            matched_header = alias_to_header.get(normalized_actual)
            if not matched_header:
                for alias_norm, canonical in alias_to_header.items():
                    if alias_norm in normalized_actual or normalized_actual in alias_norm:
                        matched_header = canonical
                        break

            if matched_header and matched_header not in header_map:
                header_map[matched_header] = col_idx

        missing_headers = [header for header in ["ADMISSION ID", "FIRST NAME", "CLASS"] if header not in header_map]
        if missing_headers:
            errors.append(
                {
                    "row": 1,
                    "error": (
                        f"Missing column(s): {', '.join(missing_headers)}. "
                        f"Found headers: {actual_headers}. Required minimum headers: "
                        "ADMISSION ID, FIRST NAME, CLASS. ROLL NO blank ho to Admission ID se auto-fill ho jayega."
                    ),
                }
            )
            return valid_students, errors

        admission_numbers_seen = set()
        roll_numbers_seen = set()

        for row_idx in range(2, worksheet.max_row + 1):
            try:
                row_values = {
                    header: worksheet.cell(row=row_idx, column=col_idx).value
                    for header, col_idx in header_map.items()
                }

                admission_id = _clean_text(row_values.get("ADMISSION ID"))
                first_name = _clean_text(row_values.get("FIRST NAME"))
                middle_name = _clean_text(row_values.get("MIDDLE NAME"))
                last_name = _clean_text(row_values.get("LAST NAME"))
                class_name = _clean_text(row_values.get("CLASS"))
                section = _clean_text(row_values.get("SECTION"))

                if not any([admission_id, first_name, last_name, class_name, section]):
                    continue

                if not admission_id or not first_name or not class_name:
                    errors.append(
                        {
                            "row": row_idx,
                            "error": "Missing required data in row. Required columns: ADMISSION ID, FIRST NAME, CLASS.",
                        }
                    )
                    continue

                normalized_admission = admission_id.lower()
                if normalized_admission in admission_numbers_seen:
                    errors.append(
                        {
                            "row": row_idx,
                            "admission_id": admission_id,
                            "error": f"Duplicate ADMISSION ID: {admission_id}.",
                        }
                    )
                    continue

                candidate_name = " ".join(part for part in [first_name, middle_name, last_name] if part).strip()
                if not candidate_name:
                    errors.append(
                        {
                            "row": row_idx,
                            "admission_id": admission_id,
                            "error": "Student name could not be derived from FIRST NAME / MIDDLE NAME / LAST NAME.",
                        }
                    )
                    continue

                managed_batch = _first_non_empty(
                    row_values.get("MANAGED BATCH"),
                    f"{class_name} | {section}" if section else class_name,
                )
                source_roll_no = _clean_text(row_values.get("ROLL NO"))
                roll_no = _first_non_empty(source_roll_no, admission_id)
                normalized_roll = roll_no.lower()
                if normalized_roll in roll_numbers_seen:
                    fallback_roll = admission_id.lower()
                    if fallback_roll and fallback_roll not in roll_numbers_seen:
                        roll_no = admission_id
                        normalized_roll = fallback_roll
                    else:
                        errors.append(
                            {
                                "row": row_idx,
                                "roll_no": roll_no,
                                "error": f"Duplicate ROLL NO after normalization: {roll_no}.",
                            }
                        )
                        continue

                boarding_type = _clean_text(row_values.get("BOARDING TYPE"))
                hostel_required = _parse_bool(row_values.get("HOSTEL REQUIRED"))
                if not hostel_required and boarding_type:
                    lowered_boarding = boarding_type.lower()
                    hostel_required = any(keyword in lowered_boarding for keyword in ["hostel", "residential", "boarder"])

                student = {
                    "sr_no": _clean_text(row_values.get("SR. NO")),
                    "admission_id": admission_id,
                    "academic_session": _first_non_empty(
                        row_values.get("ACADEMIC SESSION"),
                        default_academic_session(),
                    ),
                    "program": _clean_text(row_values.get("PROGRAM")),
                    "course": _clean_text(row_values.get("COURSE")),
                    "batch": managed_batch,
                    "managed_batch": managed_batch,
                    "class_name": class_name,
                    "section": section,
                    "roll_no": roll_no,
                    "source_roll_no": source_roll_no,
                    "first_name": first_name,
                    "middle_name": middle_name,
                    "last_name": last_name,
                    "candidate_name": candidate_name,
                    "local_name": _clean_text(row_values.get("LOCAL NAME")),
                    "dob": _parse_excel_date(row_values.get("DATE OF BIRTH")),
                    "age_as_of_today": _clean_text(row_values.get("AGE AS OF TODAY")),
                    "email": _clean_text(row_values.get("EMAIL")),
                    "gender": _clean_text(row_values.get("GENDER")),
                    "phone": _clean_text(row_values.get("PHONE")),
                    "admission_date": _parse_excel_date(row_values.get("ADMISSION DATE")),
                    "fee_schedule": _clean_text(row_values.get("FEE SCHEDULE")),
                    "tc_number": _clean_text(row_values.get("TC NUMBER")),
                    "previous_school": _clean_text(row_values.get("PREVIOUS SCHOOL")),
                    "previous_exam": _clean_text(row_values.get("PREVIOUS EXAM")),
                    "previous_board": _clean_text(row_values.get("PREVIOUS BOARD")),
                    "previous_percentage": _clean_text(row_values.get("PREVIOUS PERCENTAGE")),
                    "previous_total_marks": _clean_text(row_values.get("PREVIOUS TOTAL MARKS")),
                    "previous_average": _clean_text(row_values.get("PREVIOUS AVERAGE")),
                    "father_name": _clean_text(row_values.get("FATHER NAME")),
                    "father_mobile": _clean_text(row_values.get("FATHER MOBILE")),
                    "father_occupation": _clean_text(row_values.get("FATHER OCCUPATION")),
                    "mother_name": _clean_text(row_values.get("MOTHER NAME")),
                    "mother_mobile": _clean_text(row_values.get("MOTHER MOBILE")),
                    "mother_occupation": _clean_text(row_values.get("MOTHER OCCUPATION")),
                    "category": _clean_text(row_values.get("CATEGORY")),
                    "sub_category": _clean_text(row_values.get("SUB CATEGORY")),
                    "sibling_name": _clean_text(row_values.get("SIBLING NAME")),
                    "sibling_school": _clean_text(row_values.get("SIBLING SCHOOL")),
                    "guardian_name": _first_non_empty(row_values.get("GUARDIAN NAME"), row_values.get("FATHER NAME")),
                    "guardian_relation": _clean_text(row_values.get("GUARDIAN RELATION")),
                    "guardian_mobile": _first_non_empty(row_values.get("GUARDIAN MOBILE"), row_values.get("FATHER MOBILE")),
                    "guardian_address": _clean_text(row_values.get("GUARDIAN ADDRESS")),
                    "emergency_name": _clean_text(row_values.get("EMERGENCY NAME")),
                    "emergency_mobile": _clean_text(row_values.get("EMERGENCY MOBILE")),
                    "priority_contact": _clean_text(row_values.get("PRIORITY CONTACT")),
                    "address1": _clean_text(row_values.get("ADDRESS LINE 1")),
                    "address2": _clean_text(row_values.get("ADDRESS LINE 2")),
                    "city": _clean_text(row_values.get("CITY")),
                    "state": _clean_text(row_values.get("STATE")),
                    "country": _first_non_empty(row_values.get("COUNTRY"), "India"),
                    "pincode": _clean_text(row_values.get("PINCODE")),
                    "region": _clean_text(row_values.get("REGION")),
                    "reference_name": _clean_text(row_values.get("REFERENCE NAME")),
                    "reference_number": _clean_text(row_values.get("REFERENCE NUMBER")),
                    "reference_remark": _clean_text(row_values.get("REFERENCE REMARK")),
                    "admission_type": _clean_text(row_values.get("ADMISSION TYPE")),
                    "boarding_type": boarding_type,
                    "special_needs": _clean_text(row_values.get("SPECIAL NEEDS")),
                    "availing_mess_facility": _clean_text(row_values.get("AVAILING MESS FACILITY")),
                    "hostel_required": hostel_required,
                    "preferred_hostel": _clean_text(row_values.get("PREFERRED HOSTEL")),
                    "hostel_request_note": _clean_text(row_values.get("HOSTEL REQUEST NOTE")),
                    "pickup_enabled": _parse_bool(row_values.get("PICKUP ENABLED")),
                    "drop_enabled": _parse_bool(row_values.get("DROP ENABLED")),
                    "transport_month": _clean_text(row_values.get("TRANSPORT MONTH")),
                    "transport_route": _clean_text(row_values.get("TRANSPORT ROUTE")),
                    "transport_stop": _clean_text(row_values.get("TRANSPORT STOP")),
                    "room_no": _clean_text(row_values.get("ROOM NO")),
                }

                valid_students.append(student)
                admission_numbers_seen.add(normalized_admission)
                roll_numbers_seen.add(normalized_roll)
            except Exception as exc:
                errors.append({"row": row_idx, "error": f"Error parsing row: {str(exc)}"})
    except Exception as exc:
        import zipfile
        from openpyxl.utils.exceptions import InvalidFileException

        if isinstance(exc, (zipfile.BadZipFile, InvalidFileException)):
            errors.append({"row": 0, "error": "Invalid Excel format. The uploaded file is not a valid .xlsx file."})
        else:
            errors.append({"row": 0, "error": f"Error reading Excel file: {str(exc)}"})

    return valid_students, errors


def create_student_excel_template() -> BytesIO:
    """
    Create downloadable Excel template for student data upload.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Student Upload"
    instructions_sheet = workbook.create_sheet("Instructions")

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(STUDENT_TEMPLATE_HEADERS, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"

    sample_rows = [
        [
            1,
            "ADM-1001",
            default_academic_session(),
            "Medical",
            "NEET",
            "Foundation 7th | A",
            "Foundation 7th",
            "A",
            "ADM-1001",
            "Aarav",
            "",
            "Sharma",
            "",
            "2013-05-14",
            "13 years",
            "aarav@example.com",
            "Male",
            "9876543210",
            "2026-04-05",
            "Fees Structure",
            "",
            "Previous School Name",
            "Annual Exam",
            "CBSE",
            "87",
            "435",
            "87",
            "Rakesh Sharma",
            "9876500001",
            "Engineer",
            "Sunita Sharma",
            "9876500002",
            "Teacher",
            "General",
            "",
            "",
            "",
            "Rakesh Sharma",
            "Father",
            "9876500001",
            "Mall Road, Shimla",
            "Rakesh Sharma",
            "9876500001",
            "father",
            "House 21",
            "Near Ridge",
            "Shimla",
            "Himachal Pradesh",
            "India",
            "171001",
            "North",
            "Reference Person",
            "9800000000",
            "Known family referral",
            "new",
            "Day boarding",
            "",
            "no",
            False,
            "",
            "",
            False,
            False,
            "",
            "",
            "",
            "",
        ],
        [
            2,
            "ADM-1002",
            default_academic_session(),
            "Non Medical",
            "JEE Main",
            "Foundation 8th | B",
            "Foundation 8th",
            "B",
            "1042",
            "Priya",
            "",
            "Verma",
            "",
            "2012-08-09",
            "13 years",
            "priya@example.com",
            "Female",
            "9988776655",
            "2026-04-08",
            "Fees Structure",
            "TC-44",
            "DAV Public School",
            "Final Exam",
            "ICSE",
            "91",
            "455",
            "91",
            "Sanjay Verma",
            "9988776600",
            "Business",
            "Neha Verma",
            "9988776611",
            "Homemaker",
            "OBC",
            "",
            "",
            "",
            "Sanjay Verma",
            "Father",
            "9988776600",
            "Summer Hill",
            "Sanjay Verma",
            "9988776600",
            "father",
            "Lane 2",
            "",
            "Shimla",
            "Himachal Pradesh",
            "India",
            "171002",
            "North",
            "",
            "",
            "",
            "old",
            "Day boarding",
            "Needs front row seating",
            "no",
            False,
            "",
            "",
            False,
            False,
            "",
            "",
            "",
            "",
        ],
    ]

    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")

    instructions_sheet["A1"] = "Student Bulk Upload Instructions"
    instructions_sheet["A1"].font = Font(bold=True, size=14)
    instructions = [
        "Use only the 'Student Upload' sheet for data entry.",
        "Do not change the column headers in row 1.",
        "Required minimum fields: ADMISSION ID, FIRST NAME, CLASS.",
        "ROLL NO blank hoga to system Admission ID ko temporary roll number ki tarah use karega.",
        "CLASS + SECTION se managed batch auto-derive ho sakta hai, lekin MANAGED BATCH dena best hai.",
        "Template student form ke fields ke hisaab se aligned hai.",
        "Direct student columns ke alawa extra fields metadata me safely preserve kiye jayenge.",
        "DATE OF BIRTH aur ADMISSION DATE ko yyyy-mm-dd ya dd/mm/yyyy format me bhar sakte ho.",
        "HOSTEL REQUIRED, PICKUP ENABLED, DROP ENABLED me Yes/No ya True/False use karo.",
        "File ko .xlsx format me hi upload karo.",
        "",
        "Header order:",
        ", ".join(STUDENT_TEMPLATE_HEADERS),
    ]
    for row_idx, instruction in enumerate(instructions, 3):
        instructions_sheet.cell(row=row_idx, column=1).value = instruction

    instructions_sheet.column_dimensions["A"].width = 140

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[column].width = min(max_length + 2, 28)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_inventory_material_excel(file_content: bytes) -> Tuple[List[Dict], List[Dict]]:
    """Parse bulk inventory material rows from an Excel file."""
    valid_rows: List[Dict] = []
    errors: List[Dict] = []

    def normalize_header(value: str) -> str:
        import re
        return re.sub(r'[^a-z0-9]', '', value.strip().lower())

    alias_to_header = {
        'subjectname': 'SUBJECT NAME',
        'subject': 'SUBJECT NAME',
        'setname': 'SET NAME',
        'set': 'SET NAME',
        'target': 'SET NAME',
        'volumenumber': 'VOLUME NUMBER',
        'volumeno': 'VOLUME NUMBER',
        'volume': 'VOLUME NUMBER',
        'volumename': 'VOLUME NAME',
        'materialname': 'MATERIAL NAME',
        'material': 'MATERIAL NAME',
        'name': 'MATERIAL NAME',
        'unittype': 'UNIT TYPE',
        'unit': 'UNIT TYPE',
        'price': 'PRICE',
        'lowstockthreshold': 'LOW STOCK THRESHOLD',
        'threshold': 'LOW STOCK THRESHOLD',
        'batchnames': 'BATCH NAMES',
        'batches': 'BATCH NAMES',
        'class': 'BATCH NAMES',
        'description': 'DESCRIPTION',
        'isactive': 'IS ACTIVE',
        'active': 'IS ACTIVE',
        'suppliername': 'SUPPLIER NAME',
        'supplier': 'SUPPLIER NAME',
        'openingstock': 'OPENING STOCK',
        'openingquantity': 'OPENING STOCK',
        'quantity': 'OPENING STOCK',
        'qty': 'OPENING STOCK',
        'openingstockdate': 'STOCK IN DATE',
        'stockindate': 'STOCK IN DATE',
        'date': 'STOCK IN DATE',
    }

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        worksheet = workbook.active

        if worksheet.max_row < 2:
            return valid_rows, [{'row': 0, 'error': 'Excel file appears to be empty.'}]

        header_map: Dict[str, int] = {}
        for col_idx in range(1, worksheet.max_column + 1):
            raw_header = worksheet.cell(row=1, column=col_idx).value
            if raw_header is None:
                continue
            canonical = alias_to_header.get(normalize_header(str(raw_header)))
            if canonical and canonical not in header_map:
                header_map[canonical] = col_idx

        if 'MATERIAL NAME' not in header_map and 'DESCRIPTION' in header_map:
            header_map['MATERIAL NAME'] = header_map['DESCRIPTION']

        required_headers = ['SUBJECT NAME', 'SET NAME', 'MATERIAL NAME']
        missing = [header for header in required_headers if header not in header_map]
        if missing:
            return valid_rows, [{
                'row': 1,
                'error': f"Missing column(s): {', '.join(missing)}. Required headers: SUBJECT NAME, SET NAME, MATERIAL NAME.",
            }]

        for row_idx in range(2, worksheet.max_row + 1):
            row_data = {
                header: worksheet.cell(row=row_idx, column=col_idx).value
                for header, col_idx in header_map.items()
            }

            subject_name = str(row_data.get('SUBJECT NAME') or '').strip()
            set_name = str(row_data.get('SET NAME') or '').strip()
            material_name = str(row_data.get('MATERIAL NAME') or '').strip()

            if not subject_name and not set_name and not material_name:
                continue

            if not subject_name or not set_name or not material_name:
                errors.append({
                    'row': row_idx,
                    'error': 'SUBJECT NAME, SET NAME, and MATERIAL NAME are required.',
                })
                continue

            volume_number_raw = row_data.get('VOLUME NUMBER')
            volume_name = str(row_data.get('VOLUME NAME') or '').strip()
            unit_type = str(row_data.get('UNIT TYPE') or 'book').strip().lower()
            description = str(row_data.get('DESCRIPTION') or '').strip()
            batch_names_raw = str(row_data.get('BATCH NAMES') or '').strip()
            price_raw = row_data.get('PRICE')
            threshold_raw = row_data.get('LOW STOCK THRESHOLD')
            is_active_raw = row_data.get('IS ACTIVE')
            supplier_name = str(row_data.get('SUPPLIER NAME') or '').strip()
            stock_in_date_raw = row_data.get('STOCK IN DATE')
            opening_stock_raw = row_data.get('OPENING STOCK')

            volume_number = None
            if volume_number_raw not in (None, ''):
                try:
                    volume_number = int(volume_number_raw)
                except (TypeError, ValueError):
                    errors.append({'row': row_idx, 'error': f'Invalid VOLUME NUMBER: {volume_number_raw}'})
                    continue

            try:
                price = float(price_raw) if price_raw not in (None, '') else 0.0
            except (TypeError, ValueError):
                errors.append({'row': row_idx, 'error': f'Invalid PRICE: {price_raw}'})
                continue

            try:
                low_stock_threshold = int(threshold_raw) if threshold_raw not in (None, '') else 10
            except (TypeError, ValueError):
                errors.append({'row': row_idx, 'error': f'Invalid LOW STOCK THRESHOLD: {threshold_raw}'})
                continue

            try:
                opening_stock = int(opening_stock_raw) if opening_stock_raw not in (None, '') else 0
            except (TypeError, ValueError):
                errors.append({'row': row_idx, 'error': f'Invalid OPENING STOCK: {opening_stock_raw}'})
                continue

            if opening_stock < 0:
                errors.append({'row': row_idx, 'error': 'OPENING STOCK cannot be negative'})
                continue

            if isinstance(is_active_raw, str):
                is_active = is_active_raw.strip().lower() not in {'false', 'no', '0', 'inactive'}
            else:
                is_active = bool(is_active_raw) if is_active_raw is not None else True

            batch_names = [
                part.strip()
                for part in batch_names_raw.replace('|', ',').split(',')
                if part.strip()
            ]

            valid_rows.append({
                'subject_name': subject_name,
                'set_name': set_name,
                'volume_number': volume_number,
                'volume_name': volume_name or (f"Volume {volume_number}" if volume_number else ''),
                'material_name': material_name,
                'unit_type': unit_type or 'book',
                'price': price,
                'low_stock_threshold': low_stock_threshold,
                'batch_names': batch_names,
                'description': description,
                'is_active': is_active,
                'supplier_name': supplier_name,
                'opening_stock': opening_stock,
                'stock_in_date': stock_in_date_raw,
            })
    except Exception as e:
        import zipfile
        from openpyxl.utils.exceptions import InvalidFileException
        if isinstance(e, (zipfile.BadZipFile, InvalidFileException)):
            errors.append({'row': 0, 'error': 'Invalid Excel format. Upload a valid .xlsx file.'})
        else:
            errors.append({'row': 0, 'error': f'Error reading Excel file: {str(e)}'})

    return valid_rows, errors


def create_inventory_material_template() -> BytesIO:
    """Create downloadable Excel template for inventory material bulk upload."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Upload"
    instruction_sheet = workbook.create_sheet(title="Instructions")

    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        'TARGET',
        'DESCRIPTION',
        'SUBJECT',
        'CLASS',
        'QTY',
        'SUPPLIER NAME',
        'STOCK IN DATE',
        'LOW STOCK THRESHOLD',
        'UNIT TYPE',
        'IS ACTIVE',
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    sample_rows = [
        ['JEE (Adv)', 'Periodic Table & Periodicity', 'Chemistry', 'CLASS-XI', 50, 'ABC Supplier', '2026-04-30', 10, 'book', 'TRUE'],
        ['JEE (Adv)', 'Introduction to Chemistry', 'Chemistry', 'CLASS-XI', 50, 'ABC Supplier', '2026-04-30', 10, 'book', 'TRUE'],
        ['JEE (Adv)', 'Mole Concept', 'Chemistry', 'CLASS-XI', 50, 'ABC Supplier', '2026-04-30', 10, 'book', 'TRUE'],
        ['JEE (Adv)', 'Mathematical Tools', 'Physics', 'CLASS-XI', 50, 'ABC Supplier', '2026-04-30', 10, 'book', 'TRUE'],
        ['JEE (Adv)', 'Rectilinear Motion', 'Physics', 'CLASS-XI', 50, 'ABC Supplier', '2026-04-30', 10, 'book', 'TRUE'],
        ['JEE (Adv)', 'Trigonometry', 'Mathematics', 'CLASS-XI', 50, 'ABC Supplier', '2026-04-30', 10, 'book', 'TRUE'],
    ]
    for row_idx, row_data in enumerate(sample_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = value

    for col_letter, width in {
        'A': 18, 'B': 34, 'C': 18, 'D': 16, 'E': 10,
        'F': 22, 'G': 16, 'H': 20, 'I': 14, 'J': 12,
    }.items():
        worksheet.column_dimensions[col_letter].width = width

    instruction_sheet['A1'] = "Inventory Material Bulk Upload Template"
    instruction_sheet['A1'].font = Font(bold=True, size=14)
    instruction_sheet['A3'] = "Use this exact format: TARGET | DESCRIPTION | SUBJECT | CLASS | QTY"
    instruction_sheet['A4'] = "Mapping: TARGET -> Set Name, DESCRIPTION -> Material Name, SUBJECT -> Subject, CLASS -> Batch/Class, QTY -> Opening Stock"
    instruction_sheet['A5'] = "Optional extra columns: SUPPLIER NAME, STOCK IN DATE, LOW STOCK THRESHOLD, UNIT TYPE, IS ACTIVE"
    instruction_sheet['A6'] = "Each row is one separate inventory item. Different descriptions will stay separate."
    instruction_sheet['A7'] = "If SUPPLIER NAME and QTY diye honge, import ke time stock-in entry automatically create ho jayegi."
    instruction_sheet['A8'] = "Example image logic: JEE (Adv) + Periodic Table & Periodicity + Chemistry + CLASS-XI + 50"
    instruction_sheet['A9'] = "Aise multiple rows upload karne par supplier summary aur stock total row-wise sum hoga."
    instruction_sheet.column_dimensions['A'].width = 120

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_seating_plan_excel(file_content: bytes) -> Tuple[List[Dict], List[Dict]]:
    valid_entries = []
    errors = []

    def normalize_header(value: str) -> str:
        import re
        return re.sub(r'[^a-z0-9]', '', value.strip().lower())

    required_headers = ['ROLL NO', 'CANDIDATE NAME', 'BATCH', 'ROOM NO']
    optional_headers = ['SR. NO', 'FATHER NAME']
    header_aliases = {
        'SR. NO': ['srno', 'sno', 's.no', 'serialno', 'serial', 'serialnumber', 'sr'],
        'ROLL NO': ['rollno', 'rollnumber', 'roll number', 'roll', 'registrationno', 'registrationnumber', 'roll#', 'rollno.'],
        'CANDIDATE NAME': ['candidatename', 'studentname', 'student name', 'name', 'candidate', 'fullname', 'full name'],
        'FATHER NAME': ['fathername', 'father name', 'father', 'parentname', 'parent name', 'guardianname', 'guardian name', 'guardian'],
        'BATCH': ['batch', 'class', 'standard', 'grade', 'year'],
        'ROOM NO': ['roomno', 'room number', 'room', 'classroom', 'room#', 'class room', 'roomno.'],
    }

    alias_to_header = {}
    for canonical, aliases in header_aliases.items():
        alias_to_header[normalize_header(canonical)] = canonical
        for alias in aliases:
            alias_to_header[normalize_header(alias)] = canonical

    def find_header_row() -> Tuple[int, Dict[str, int], List[str]]:
        max_scan = min(5, worksheet.max_row)
        max_columns = min(15, worksheet.max_column)

        for row_idx in range(1, max_scan + 1):
            actual_headers = [
                str(worksheet.cell(row=row_idx, column=col_idx).value).strip()
                if worksheet.cell(row=row_idx, column=col_idx).value is not None else ''
                for col_idx in range(1, max_columns + 1)
            ]

            normalized_headers = [normalize_header(value) for value in actual_headers]
            header_map: Dict[str, int] = {}

            for col_idx, normalized_actual in enumerate(normalized_headers, start=1):
                canonical = None
                if normalized_actual in alias_to_header:
                    canonical = alias_to_header[normalized_actual]
                else:
                    for alias_norm, alias_header in alias_to_header.items():
                        if alias_norm in normalized_actual or normalized_actual in alias_norm:
                            canonical = alias_header
                            break
                if canonical and canonical not in header_map:
                    header_map[canonical] = col_idx

            if all(header in header_map for header in required_headers):
                return row_idx, header_map, actual_headers

        return 0, {}, []

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        worksheet = workbook.active

        if worksheet.max_row < 2:
            errors.append({'row': 0, 'error': 'Excel file appears to be empty.'})
            return valid_entries, errors

        header_row, header_map, _actual_headers = find_header_row()
        if not header_map:
            errors.append({
                'row': 1,
                'error': f'Invalid Excel format. Missing required columns. Expected headers: {required_headers + optional_headers}. Found headers in first rows.'
            })
            return valid_entries, errors

        for optional_header in optional_headers:
            if optional_header not in header_map:
                header_map[optional_header] = None

        roll_numbers_seen = set()

        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            try:
                def get_cell_value(header_name: str):
                    col = header_map.get(header_name)
                    if col is None:
                        return None
                    return worksheet.cell(row=row_idx, column=col).value

                sr_no = get_cell_value('SR. NO')
                roll_no = get_cell_value('ROLL NO')
                candidate_name = get_cell_value('CANDIDATE NAME')
                father_name = get_cell_value('FATHER NAME')
                batch = get_cell_value('BATCH')
                room_no = get_cell_value('ROOM NO')

                if not roll_no or not candidate_name or not batch or not room_no:
                    errors.append({'row': row_idx, 'error': 'Missing required data (ROLL NO, CANDIDATE NAME, BATCH, ROOM NO).'})
                    continue

                roll_no_str = str(roll_no).strip()
                if not roll_no_str:
                    errors.append({'row': row_idx, 'error': 'ROLL NO cannot be empty.'})
                    continue

                if roll_no_str in roll_numbers_seen:
                    errors.append({'row': row_idx, 'error': f'Duplicate ROLL NO: {roll_no_str}.'})
                    continue

                roll_numbers_seen.add(roll_no_str)

                batch_str = str(batch).strip()
                if not batch_str:
                    errors.append({'row': row_idx, 'error': 'BATCH cannot be empty.'})
                    continue

                valid_entries.append({
                    'sr_no': sr_no,
                    'roll_no': roll_no_str,
                    'candidate_name': str(candidate_name).strip(),
                    'father_name': str(father_name).strip() if father_name else '',
                    'batch': batch_str,
                    'room_no': str(room_no).strip(),
                })
            except Exception as e:
                errors.append({'row': row_idx, 'error': f'Error parsing row: {str(e)}'})
    except Exception as e:
        import zipfile
        from openpyxl.utils.exceptions import InvalidFileException
        if isinstance(e, (zipfile.BadZipFile, InvalidFileException)):
            errors.append({'row': 0, 'error': 'Invalid Excel format. The uploaded file is not a valid .xlsx file.'})
        else:
            errors.append({'row': 0, 'error': f'Error reading Excel file: {str(e)}'})

    return valid_entries, errors


def create_seating_plan_template() -> BytesIO:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Seating Plan Template"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    worksheet['A1'] = "ASPIRE IIT & MEDICAL - Seating Plan Upload Template"
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A2'] = "IMPORTANT: Do not modify column headers. Fill data starting from row 3."
    worksheet['A2'].font = Font(bold=True, color="FF6600")
    worksheet['A3'] = "Required fields are marked with *"
    worksheet['A3'].font = Font(italic=True)

    headers = ['SR. NO*', 'ROLL NO*', 'CANDIDATE NAME*', 'FATHER NAME', 'BATCH*', 'ROOM NO*']
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=5, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    sample_rows = [
        [1, '101', 'Rahul Sharma', 'Rakesh Sharma', '12th Medical', 'Room 1'],
        [2, '102', 'Priya Patel', 'Rajesh Patel', '12th IIT', 'Room 1'],
        [3, '103', 'Amit Kumar', 'Suresh Kumar', '12th Medical', 'Room 2'],
        [4, '104', 'Sneha Singh', 'Vikram Singh', 'Dropper 1', 'Room 2'],
    ]
    for row_idx, row_data in enumerate(sample_rows, 6):
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center')

    instructions_start_row = 10
    worksheet.cell(row=instructions_start_row, column=1).value = "INSTRUCTIONS:"
    worksheet.cell(row=instructions_start_row, column=1).font = Font(bold=True)
    instructions = [
        "1. SR. NO: Sequential number (1, 2, 3, ...)",
        "2. ROLL NO: Unique student roll number (required)",
        "3. CANDIDATE NAME: Full name of the student (required)",
        "4. FATHER NAME: Father's name (optional)",
        "5. BATCH: Any batch name is allowed and will be created automatically during import",
        "6. ROOM NO: Room assignment (e.g., 'Room 1', 'Room 2', etc.)",
        "",
        "NOTES:",
        "• All students with same ROOM NO will be grouped together",
        "• ROLL NO must be unique across the entire file",
        "• Save file as .xlsx format only",
        "• New batch names from Excel will be auto-added to the system",
    ]
    for idx, instruction in enumerate(instructions):
        worksheet.cell(row=instructions_start_row + 1 + idx, column=1).value = instruction

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[column].width = min(max_length + 2, 50)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def create_seating_export_excel(plan_data: Dict, room_data: Dict) -> BytesIO:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Seating Plan"

    _build_roomwise_seating_sheet(
        worksheet,
        [{
            "plan_data": plan_data,
            "room_data": room_data,
        }],
        title_override=room_data.get("exam_name") or "SITTING PLAN",
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def create_multi_room_seating_export_excel(room_plans: List[Dict]) -> BytesIO:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "All Rooms Seating"
    summary_sheet = workbook.create_sheet(title="Room Summary")

    title_override = ""
    if room_plans:
        first_room = room_plans[0].get("room_data") or {}
        title_override = first_room.get("exam_name") or "SITTING PLAN"

    _ensure_room_students_cache(room_plans)
    _build_roomwise_seating_sheet(worksheet, room_plans, title_override=title_override)
    _build_room_summary_sheet(summary_sheet, room_plans, title_override=title_override)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _normalize_room_students(plan_data: Dict) -> List[Dict]:
    normalized_rows: List[Dict] = []
    sequence = 1
    for desk_id, students in plan_data.get("assignment", {}).items():
        try:
            desk_sort = int(desk_id)
        except Exception:
            desk_sort = sequence

        for seat_pos, student in enumerate(students, 1):
            normalized_rows.append({
                "sequence": sequence,
                "desk_sort": desk_sort,
                "seat_pos": seat_pos,
                "roll_number": student.get("roll_number", ""),
                "name": student.get("name", ""),
                "father_name": student.get("father_name", ""),
                "batch": student.get("batch", ""),
            })
            sequence += 1

    normalized_rows.sort(key=lambda item: (item["desk_sort"], item["seat_pos"], item["sequence"]))
    for index, row in enumerate(normalized_rows, start=1):
        row["sequence"] = index
    return normalized_rows


def _ensure_room_students_cache(room_plans: List[Dict]) -> None:
    for room_plan in room_plans:
        if "_normalized_students" in room_plan:
            continue
        plan_data = room_plan.get("plan_data") or {}
        room_plan["_normalized_students"] = _normalize_room_students(plan_data)


def _build_roomwise_seating_sheet(worksheet, room_plans: List[Dict], title_override: str = "") -> None:
    title_fill = PatternFill(start_color="F4B4A8", end_color="F4B4A8", fill_type="solid")
    room_fill = PatternFill(start_color="F4B4A8", end_color="F4B4A8", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    white_bold_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    room_font = Font(bold=True, size=13)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    headers = ["Sr.No.", "Roll Number", "Student Name", "Father name", "Batch", "Room"]
    title_text = f"SITTING PLAN FOR {str(title_override or 'EXAM').upper()}"
    if room_plans and (room_plans[0].get("room_data") or {}).get("exam_subject"):
        exam_subject = str((room_plans[0].get("room_data") or {}).get("exam_subject") or "").strip()
        if exam_subject:
            title_text = f"SITTING PLAN FOR {exam_subject.upper()}"

    worksheet.merge_cells("A1:F1")
    title_cell = worksheet["A1"]
    title_cell.value = title_text
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = title_fill

    current_row = 2
    for room_plan in room_plans:
        plan_data = room_plan.get("plan_data") or {}
        room_data = room_plan.get("room_data") or {}
        room_name = str(room_data.get("name") or "").strip() or "ROOM"

        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        room_cell = worksheet.cell(row=current_row, column=1)
        room_cell.value = f"ROOM NO. - {room_name}"
        room_cell.font = room_font
        room_cell.alignment = Alignment(horizontal="center", vertical="center")
        room_cell.fill = room_fill

        current_row += 1
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        current_row += 1
        room_students = room_plan.get("_normalized_students") or _normalize_room_students(plan_data)
        for row in room_students:
            values = [
                row["sequence"],
                row["roll_number"],
                row["name"],
                row["father_name"],
                row["batch"],
                room_name,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            current_row += 1

        current_row += 1

    for merge_range in list(worksheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merge_range.bounds
        for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = thin_border

    widths = {"A": 10, "B": 26, "C": 30, "D": 26, "E": 30, "F": 14}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _build_room_summary_sheet(worksheet, room_plans: List[Dict], title_override: str = "") -> None:
    title_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_fill = PatternFill(start_color="EAF3FF", end_color="EAF3FF", fill_type="solid")
    white_bold_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    room_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    worksheet.merge_cells("A1:E1")
    title_cell = worksheet["A1"]
    title_cell.value = f"ROOM-WISE SUMMARY FOR {str(title_override or 'EXAM').upper()}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = title_fill
    title_cell.border = thin_border

    current_row = 3
    for room_plan in room_plans:
        room_data = room_plan.get("room_data") or {}
        plan_data = room_plan.get("plan_data") or {}
        room_name = str(room_data.get("name") or "").strip() or "ROOM"
        room_students = room_plan.get("_normalized_students") or _normalize_room_students(plan_data)

        batch_counts: Dict[str, int] = {}
        for student in room_students:
            batch_name = str(student.get("batch") or "").strip() or "Unassigned"
            batch_counts[batch_name] = batch_counts.get(batch_name, 0) + 1

        sorted_batches = sorted(batch_counts.items(), key=lambda item: (-item[1], item[0].lower()))

        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        room_cell = worksheet.cell(row=current_row, column=1)
        room_cell.value = f"ROOM NO. - {room_name} | TOTAL STUDENTS: {len(room_students)} | TOTAL BATCHES: {len(sorted_batches)}"
        room_cell.font = room_font
        room_cell.alignment = Alignment(horizontal="left", vertical="center")
        room_cell.fill = section_fill
        room_cell.border = thin_border
        current_row += 1

        headers = ["Sr.No.", "Batch Name", "Student Count", "Share %", "Room"]
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        current_row += 1

        total_students = len(room_students)
        for index, (batch_name, count) in enumerate(sorted_batches, start=1):
            share = round((count / total_students) * 100, 2) if total_students else 0
            values = [index, batch_name, count, share, room_name]
            for col_idx, value in enumerate(values, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            current_row += 1

        current_row += 1

    for merge_range in list(worksheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merge_range.bounds
        for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = thin_border

    widths = {"A": 10, "B": 32, "C": 16, "D": 12, "E": 18}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

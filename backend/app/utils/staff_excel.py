from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook import Workbook

STAFF_TEMPLATE_HEADERS = [
    "STAFF TYPE",
    "STAFF CATEGORY",
    "FIRST NAME",
    "MIDDLE NAME",
    "LAST NAME",
    "DOB",
    "PRIMARY MOBILE",
    "WHATSAPP NUMBER",
    "EMAIL",
    "GENDER",
    "MARITAL STATUS",
    "EMPLOYEE ID",
    "JOINING DATE",
    "SUBJECT",
    "DEPARTMENT",
    "DESIGNATION",
    "SHIFT TIMING",
    "SCHOOL NAME",
    "FATHER NAME",
    "FATHER CONTACT",
    "MOTHER NAME",
    "MOTHER CONTACT",
    "SPOUSE NAME",
    "ADDRESS LINE 1",
    "ADDRESS LINE 2",
    "CITY",
    "STATE",
    "COUNTRY",
    "PIN CODE",
    "AADHAAR NUMBER",
    "PAN NUMBER",
    "BLOOD GROUP",
    "EMERGENCY CONTACT NAME",
    "EMERGENCY CONTACT NUMBER",
    "EMERGENCY RELATION",
    "MONTHLY SALARY",
    "ACCOUNT NUMBER",
    "IFSC CODE",
    "BANK NAME",
    "NOTES",
    "IS ACTIVE",
]


def _normalize(value: Any) -> str:
    return str(value or "").strip()


def _normalize_header(value: str) -> str:
    return "".join(char.lower() for char in str(value or "").strip() if char.isalnum())


def _normalize_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _normalize(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = _normalize(value).lower()
    if text in {"false", "no", "0", "inactive"}:
        return False
    if text in {"true", "yes", "1", "active"}:
        return True
    return default


def _split_name_parts(full_name: str, explicit_last_name: str = "") -> tuple[str, str, str]:
    full_name = _normalize(full_name)
    explicit_last_name = _normalize(explicit_last_name)
    if explicit_last_name:
        base = full_name.removesuffix(explicit_last_name).strip() if full_name.lower().endswith(explicit_last_name.lower()) else full_name
        parts = [part for part in base.split() if part]
        if not parts:
            return explicit_last_name, "", ""
        if len(parts) == 1:
            return parts[0], "", explicit_last_name
        return parts[0], " ".join(parts[1:]), explicit_last_name
    parts = [part for part in full_name.split() if part]
    if not parts:
        return "", "", ""
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[0], "", parts[1]
    return parts[0], " ".join(parts[1:-1]), parts[-1]


SOURCE_HEADER_ALIASES = {
    "stafftype": "STAFF TYPE",
    "staffcategory": "STAFF CATEGORY",
    "firstname": "FIRST NAME",
    "middlename": "MIDDLE NAME",
    "lastname": "LAST NAME",
    "dob": "DOB",
    "primarymobile": "PRIMARY MOBILE",
    "phonenumber": "PRIMARY MOBILE",
    "whatsappnumber": "WHATSAPP NUMBER",
    "whatsappphonenumber": "WHATSAPP NUMBER",
    "email": "EMAIL",
    "gender": "GENDER",
    "maritalstatus": "MARITAL STATUS",
    "employeeid": "EMPLOYEE ID",
    "joiningdate": "JOINING DATE",
    "subject": "SUBJECT",
    "department": "DEPARTMENT",
    "designation": "DESIGNATION",
    "shifttiming": "SHIFT TIMING",
    "schoolname": "SCHOOL NAME",
    "fathername": "FATHER NAME",
    "fathercontact": "FATHER CONTACT",
    "fathercontactnumber": "FATHER CONTACT",
    "fatherhusbandname": "FATHER NAME",
    "mothername": "MOTHER NAME",
    "mothercontact": "MOTHER CONTACT",
    "mothercontactnumber": "MOTHER CONTACT",
    "spousename": "SPOUSE NAME",
    "addressline1": "ADDRESS LINE 1",
    "housenumber": "ADDRESS LINE 1",
    "addressline2": "ADDRESS LINE 2",
    "location": "ADDRESS LINE 2",
    "city": "CITY",
    "state": "STATE",
    "country": "COUNTRY",
    "zipcode": "PIN CODE",
    "zip": "PIN CODE",
    "pincode": "PIN CODE",
    "adharnumber": "AADHAAR NUMBER",
    "aadhaarnumber": "AADHAAR NUMBER",
    "pannumber": "PAN NUMBER",
    "bloodgroup": "BLOOD GROUP",
    "emergencycontactname": "EMERGENCY CONTACT NAME",
    "emergencycontactperson": "EMERGENCY CONTACT NAME",
    "emergencycontactnumber": "EMERGENCY CONTACT NUMBER",
    "relationshipwithemergencycontact": "EMERGENCY RELATION",
    "emergencyrelation": "EMERGENCY RELATION",
    "monthlysalary": "MONTHLY SALARY",
    "accountnumber": "ACCOUNT NUMBER",
    "ifsccode": "IFSC CODE",
    "bankname": "BANK NAME",
    "notes": "NOTES",
    "medicalandhealthcondition": "NOTES",
    "isactive": "IS ACTIVE",
}


EXTRA_SOURCE_FIELDS = {
    "race",
    "religion",
    "biometricid",
    "highestqualification",
    "yearsofservice",
    "trainingrecord",
    "educationworkerpermit",
    "reportingto",
    "linkedschools",
    "examinationincharge",
    "coordinator",
    "numberofchildren",
    "childname",
    "childdobddmmyyyy",
    "childgender",
    "childage",
    "nationality",
    "identificationpassportnumber",
    "passportexpirydateddmmyyyy",
    "accountname",
    "nextincrementdueddmmyyyy",
    "vehicleinformation",
    "drivingyesno",
}


def create_staff_excel_template() -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Staff Upload"
    instructions_sheet = workbook.create_sheet("Instructions")

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(STAFF_TEMPLATE_HEADERS, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_rows = [
        [
            "teaching", "Teacher", "Aman", "", "Sharma", "1990-04-12", "9876543210", "9876543210",
            "aman@example.com", "Male", "Single", "EMP-T-001", "2026-04-01", "Mathematics", "Mathematics",
            "Teacher", "8 AM - 3 PM", "Dr. Girish App", "Ramesh Sharma", "9988776655", "", "", "",
            "House 12", "Near Market", "Shimla", "Himachal Pradesh", "India", "171001", "123412341234",
            "ABCDE1234F", "O+", "Ramesh Sharma", "9988776655", "Father", "45000", "1234567890",
            "SBIN0001234", "State Bank of India", "Demo teaching row", "true",
        ],
        [
            "non_teaching", "Office Staff", "Rakesh", "", "Kumar", "1988-02-03", "9876500000", "",
            "rakesh@example.com", "Male", "Married", "EMP-NT-001", "2026-04-03", "", "Operations",
            "Office Staff", "9 AM - 5 PM", "Dr. Girish App", "", "", "", "", "",
            "Lane 3", "", "Shimla", "Himachal Pradesh", "India", "171002", "", "",
            "A+", "Seema Kumari", "9876500001", "Spouse", "28000", "998877665544",
            "PUNB0123456", "Punjab National Bank", "Demo non-teaching row", "true",
        ],
    ]

    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")

    instructions_sheet["A1"] = "Staff Bulk Upload Instructions"
    instructions_sheet["A1"].font = Font(bold=True, size=14)
    instructions = [
        "Use only the 'Staff Upload' sheet for data entry.",
        "Do not change row 1 column headers.",
        "Required fields: STAFF TYPE, FIRST NAME, EMPLOYEE ID.",
        "Teaching staff ke liye SUBJECT required hai.",
        "Non-teaching staff ke liye DEPARTMENT ya DESIGNATION dena best rahega.",
        "DATE fields ko yyyy-mm-dd ya dd/mm/yyyy me bhar sakte ho.",
        "This template Add Staff form ke fields ke hisaab se aligned hai.",
        "Extra source-specific columns upload ke waqt metadata me preserve kiye ja sakte hain.",
        "",
        "Header order:",
        ", ".join(STAFF_TEMPLATE_HEADERS),
    ]
    for row_idx, instruction in enumerate(instructions, 3):
        instructions_sheet.cell(row=row_idx, column=1).value = instruction
    instructions_sheet.column_dimensions["A"].width = 140

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except Exception:
                pass
        worksheet.column_dimensions[column].width = min(max_length + 2, 28)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_staff_excel(file_content: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    valid_rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    workbook = openpyxl.load_workbook(BytesIO(file_content))
    worksheet = workbook.active
    if worksheet.max_row < 2:
        return valid_rows, [{"row": 0, "error": "Excel file appears to be empty."}]

    header_map: Dict[str, int] = {}
    raw_header_by_canonical: Dict[str, str] = {}
    extra_header_map: Dict[str, int] = {}

    for col_idx in range(1, worksheet.max_column + 1):
        raw_header = worksheet.cell(row=1, column=col_idx).value
        if raw_header is None:
            continue
        raw_header_text = str(raw_header).strip()
        normalized = _normalize_header(raw_header_text.replace("*", ""))
        canonical = SOURCE_HEADER_ALIASES.get(normalized, raw_header_text.upper() if raw_header_text.upper() in STAFF_TEMPLATE_HEADERS else None)
        if canonical and canonical not in header_map:
            header_map[canonical] = col_idx
            raw_header_by_canonical[canonical] = raw_header_text
        elif normalized in EXTRA_SOURCE_FIELDS:
            extra_header_map[raw_header_text] = col_idx

    required = ["STAFF TYPE", "FIRST NAME", "EMPLOYEE ID"]
    missing = [header for header in required if header not in header_map]
    if missing:
        return valid_rows, [{"row": 1, "error": f"Missing column(s): {', '.join(missing)}"}]

    for row_idx in range(2, worksheet.max_row + 1):
        row_data = {
            header: worksheet.cell(row=row_idx, column=col_idx).value
            for header, col_idx in header_map.items()
        }
        extra_values = {
            raw_header: worksheet.cell(row=row_idx, column=col_idx).value
            for raw_header, col_idx in extra_header_map.items()
            if worksheet.cell(row=row_idx, column=col_idx).value not in (None, "")
        }

        if not any(_normalize(value) for value in row_data.values()):
            continue

        source_first_name = _normalize(row_data.get("FIRST NAME"))
        first_name, middle_name, last_name = _split_name_parts(
            source_first_name,
            _normalize(row_data.get("LAST NAME")),
        )
        if not first_name:
            errors.append({"row": row_idx, "error": "FIRST NAME is required."})
            continue

        staff_type_raw = _normalize(row_data.get("STAFF TYPE")).lower().replace("-", "_").replace(" ", "_")
        if staff_type_raw in {"teaching", "teacher"}:
            staff_type = "teaching"
        elif staff_type_raw in {"non_teaching", "nonteaching", "invigilator", "staff"}:
            staff_type = "non_teaching"
        else:
            errors.append({"row": row_idx, "error": f"Invalid STAFF TYPE: {_normalize(row_data.get('STAFF TYPE'))}"})
            continue

        employee_id = _normalize(row_data.get("EMPLOYEE ID"))
        if not employee_id:
            errors.append({"row": row_idx, "error": "EMPLOYEE ID is required."})
            continue

        department = _normalize(row_data.get("DEPARTMENT"))
        designation = _normalize(row_data.get("DESIGNATION"))
        subject = _normalize(row_data.get("SUBJECT")) or (department if staff_type == "teaching" else "")
        if staff_type == "teaching" and not subject:
            errors.append({"row": row_idx, "employee_id": employee_id, "error": "Teaching staff requires SUBJECT or DEPARTMENT."})
            continue

        category = _normalize(row_data.get("STAFF CATEGORY")) or (
            "Teacher" if staff_type == "teaching" else designation or department or "Non-Teaching Staff"
        )
        notes_parts = [_normalize(row_data.get("NOTES"))]
        for raw_header, value in extra_values.items():
            text = _normalize(value)
            if text:
                notes_parts.append(f"{raw_header}: {text}")

        valid_rows.append(
            {
                "staff_type": staff_type,
                "staff_category": category,
                "first_name": first_name,
                "middle_name": middle_name,
                "last_name": last_name,
                "dob": _normalize_date(row_data.get("DOB")),
                "primary_mobile": _normalize(row_data.get("PRIMARY MOBILE")),
                "whatsapp_number": _normalize(row_data.get("WHATSAPP NUMBER")),
                "email": _normalize(row_data.get("EMAIL")),
                "gender": _normalize(row_data.get("GENDER")),
                "marital_status": _normalize(row_data.get("MARITAL STATUS")),
                "employee_id": employee_id,
                "joining_date": _normalize_date(row_data.get("JOINING DATE")),
                "subject": subject,
                "department": department or (subject if staff_type == "teaching" else ""),
                "designation": designation or category,
                "shift_timing": _normalize(row_data.get("SHIFT TIMING")),
                "school_name": _normalize(row_data.get("SCHOOL NAME")),
                "father_name": _normalize(row_data.get("FATHER NAME")),
                "father_contact": _normalize(row_data.get("FATHER CONTACT")),
                "mother_name": _normalize(row_data.get("MOTHER NAME")),
                "mother_contact": _normalize(row_data.get("MOTHER CONTACT")),
                "spouse_name": _normalize(row_data.get("SPOUSE NAME")),
                "address_line_1": _normalize(row_data.get("ADDRESS LINE 1")),
                "address_line_2": _normalize(row_data.get("ADDRESS LINE 2")),
                "city": _normalize(row_data.get("CITY")),
                "state": _normalize(row_data.get("STATE")),
                "country": _normalize(row_data.get("COUNTRY")) or "India",
                "pin_code": _normalize(row_data.get("PIN CODE")),
                "aadhaar_number": _normalize(row_data.get("AADHAAR NUMBER")),
                "pan_number": _normalize(row_data.get("PAN NUMBER")),
                "blood_group": _normalize(row_data.get("BLOOD GROUP")),
                "emergency_contact_name": _normalize(row_data.get("EMERGENCY CONTACT NAME")),
                "emergency_contact_number": _normalize(row_data.get("EMERGENCY CONTACT NUMBER")),
                "emergency_relation": _normalize(row_data.get("EMERGENCY RELATION")),
                "monthly_salary": _normalize(row_data.get("MONTHLY SALARY")),
                "account_number": _normalize(row_data.get("ACCOUNT NUMBER")),
                "ifsc_code": _normalize(row_data.get("IFSC CODE")),
                "bank_name": _normalize(row_data.get("BANK NAME")),
                "notes": "\n".join(part for part in notes_parts if part),
                "is_active": _parse_bool(row_data.get("IS ACTIVE"), default=True),
                "source_headers": raw_header_by_canonical,
                "source_extras": extra_values,
            }
        )

    return valid_rows, errors

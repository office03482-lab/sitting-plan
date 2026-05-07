#!/usr/bin/env python3
"""Test Excel parsing functionality"""

from app.utils.excel import create_student_excel_template, parse_student_excel, parse_seating_plan_excel
import openpyxl
from io import BytesIO

def test_excel_parsing():
    print("Testing Excel parsing...")

    # Test 1: Parse the template
    template = create_student_excel_template()
    valid_students, errors = parse_student_excel(template.getvalue())
    print(f"Template parsing: {len(valid_students)} students, {len(errors)} errors")

    # Test 2: Create a test file with different headers
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    # Different header format
    worksheet['A1'] = 'S.No'
    worksheet['B1'] = 'Roll Number'
    worksheet['C1'] = 'Student Name'
    worksheet['D1'] = 'Father'
    worksheet['E1'] = 'Class'
    worksheet['F1'] = 'Room'
    
    # Sample data
    worksheet['A2'] = 1
    worksheet['B2'] = 'SSB001'
    worksheet['C2'] = 'John Doe'
    worksheet['D2'] = 'Mr. Doe'
    worksheet['E2'] = '12th Medical'
    worksheet['F2'] = 'Room A'
    
    worksheet['A3'] = 2
    worksheet['B3'] = 'SSB002'
    worksheet['C3'] = 'Jane Smith'
    worksheet['D3'] = 'Mr. Smith'
    worksheet['E3'] = 'Dropper 1'
    worksheet['F3'] = 'Room B'
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # Test parsing
    valid_students2, errors2 = parse_student_excel(output.getvalue())
    print(f"Alternative headers: {len(valid_students2)} students, {len(errors2)} errors")
    
    if valid_students2:
        print(f"  Sample: {valid_students2[0]}")

    print("All tests completed!")


def test_seating_excel_parsing_with_headers_on_row_3():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'Seating plan upload template'
    worksheet['A2'] = 'Please fill data starting from row 3'
    worksheet['A3'] = 'Roll Number'
    worksheet['B3'] = 'Student Name'
    worksheet['C3'] = 'Class'
    worksheet['D3'] = 'Room'

    worksheet['A4'] = 'R001'
    worksheet['B4'] = 'Alice Smith'
    worksheet['C4'] = '12th'
    worksheet['D4'] = 'Room 101'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    valid_entries, errors = parse_seating_plan_excel(output.getvalue())
    print(f"Seating parser row 3 header detection: {len(valid_entries)} valid entries, {len(errors)} errors")
    assert len(valid_entries) == 1
    assert not errors


def test_seating_excel_parsing_with_optional_headers():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'Roll Number'
    worksheet['B1'] = 'Student Name'
    worksheet['C1'] = 'Class'
    worksheet['D1'] = 'Room'

    worksheet['A2'] = 'R002'
    worksheet['B2'] = 'Bob Johnson'
    worksheet['C2'] = '11th'
    worksheet['D2'] = 'Room 102'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    valid_entries, errors = parse_seating_plan_excel(output.getvalue())
    print(f"Seating parser optional headers: {len(valid_entries)} valid entries, {len(errors)} errors")
    assert len(valid_entries) == 1
    assert not errors


if __name__ == "__main__":
    test_excel_parsing()
    test_seating_excel_parsing_with_headers_on_row_3()
    test_seating_excel_parsing_with_optional_headers()